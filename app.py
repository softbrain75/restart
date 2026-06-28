from __future__ import annotations

import io
import json
import hmac
import math
import os
import re
import uuid
from copy import deepcopy
from datetime import UTC, date, datetime, timedelta, timezone
from functools import wraps
from pathlib import Path
from typing import Any

import xlrd
from calculation import calculate_case_result
from flask import Flask, redirect, render_template, request, send_file, session, url_for
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from itsdangerous import BadSignature, SignatureExpired, URLSafeTimedSerializer
from werkzeug.security import check_password_hash, generate_password_hash


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
USERS_PATH = DATA_DIR / "users.json"
CASES_DIR = DATA_DIR / "cases"
LEADS_DIR = BASE_DIR / "leads"
CONSULTATION_LOG_PATH = LEADS_DIR / "consultations.jsonl"
ALLOWED_EXTENSIONS = {".xls", ".xlsx"}
WHITESPACE_RE = re.compile(r"[\s\u00a0\u200b\u200c\u200d\ufeff]+")
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
DOCUMENT_FIELDS = (
    ("balance_file", "재무제표"),
    ("income_file", "손익계산서"),
)
EMAIL_VERIFICATION_MAX_AGE_SECONDS = 60 * 60 * 24 * 7
EMAIL_VERIFICATION_SALT = "restart-email-verification"
ADMIN_STATUS_OPTIONS = (
    ("new", "신규"),
    ("reviewing", "검토중"),
    ("contacted", "연락완료"),
    ("scheduled", "상담예약"),
    ("hold", "보류"),
    ("closed", "종료"),
)
ADMIN_STATUS_LABELS = dict(ADMIN_STATUS_OPTIONS)
ADMIN_STATUS_TONES = {
    "new": "warning",
    "reviewing": "neutral",
    "contacted": "positive",
    "scheduled": "positive",
    "hold": "negative",
    "closed": "neutral",
}

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 40 * 1024 * 1024
app.secret_key = os.environ.get("RESTART_SECRET_KEY", "restart-local-dev-secret")
KST = timezone(timedelta(hours=9), "KST")


@app.after_request
def prevent_html_cache(response):
    if response.content_type.startswith("text/html"):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response

def now_iso() -> str:
    return datetime.now(UTC).isoformat(timespec="seconds")


def format_kst_datetime(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return text
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=UTC)
    return parsed.astimezone(KST).strftime("%Y-%m-%d %H:%M:%S")


def read_json_file(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        with path.open("r", encoding="utf-8") as file:
            return json.load(file)
    except (OSError, json.JSONDecodeError):
        return default


def write_json_file(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_suffix(path.suffix + ".tmp")
    with temp_path.open("w", encoding="utf-8") as file:
        json.dump(payload, file, ensure_ascii=False, indent=2)
    temp_path.replace(path)


def load_users() -> dict[str, dict[str, Any]]:
    users = read_json_file(USERS_PATH, {})
    return users if isinstance(users, dict) else {}


def save_users() -> None:
    write_json_file(USERS_PATH, USERS)


def load_cases() -> dict[str, dict[str, Any]]:
    cases: dict[str, dict[str, Any]] = {}
    if not CASES_DIR.exists():
        return cases
    for path in CASES_DIR.glob("*.json"):
        case = read_json_file(path, None)
        if isinstance(case, dict) and case.get("case_id"):
            cases[case["case_id"]] = case
    return cases


def write_case(case: dict[str, Any]) -> None:
    if not case.get("case_id"):
        return
    write_json_file(CASES_DIR / f"{case['case_id']}.json", case)


def save_case(case: dict[str, Any]) -> None:
    if not case.get("case_id"):
        return
    case.setdefault("created_at", now_iso())
    case["updated_at"] = now_iso()
    write_case(case)


USERS: dict[str, dict[str, Any]] = load_users()
CASE_STORE: dict[str, dict[str, Any]] = load_cases()
NON_EDITABLE_ACCOUNT_RE = re.compile(
    r"^(?:[\dIVXLCDMivxlcdm\u2160-\u217F\u2460-\u24FF]|"
    r"[\(\[（［]\s*(?:\d+|[IVXLCDMivxlcdm]+|[\u2160-\u217F]+|[\u2460-\u24FF]|[가-힣])\s*[\)\]）］])"
)
ACCOUNT_CODE_RE = re.compile(r"^\[(?P<code>\d{4,})\]")
PREPAID_ACCOUNTS = {"선급금", "선급비용"}
ASSET_TOTAL_ACCOUNTS = {"자산총계"}
LIABILITY_SECTION_START_ACCOUNTS = {"부채", "부채및자본"}
DEDUCTIBLE_ASSET_ACCOUNTS = ("감가상각누계", "국고보조", "대손충당금")
INTANGIBLE_ASSET_ACCOUNT = "무형자산"
DEBT_DEFAULT_ROWS = (
    ("secured_debt", "담보채무"),
    ("unsecured_financial_debt", "무담보 금융기관채무"),
    ("other_unsecured_debt", "기타 무담보채무(상거래채무 등)"),
    ("related_party_debt", "특수관계인채무"),
    ("unpaid_wages", "미지급급여, 미지급퇴직금(세후)"),
    ("retirement_benefit", "퇴직급여추계액"),
    ("tax_arrears", "조세체납금액(4대보험체납금액 포함)"),
)
COLLATERAL_DEFAULT_ROWS = (
    ("collateral_except_machinery", "담보제공자산(기계장치 제외)"),
    ("collateral_machinery", "담보제공 기계장치"),
    ("savings", "정기예.적금"),
    ("insurance", "보험해약환급금"),
    ("securities", "유가증권"),
    ("other_non_business_assets", "기타 비업무용 자산"),
)
RENT_DEFAULT_ROWS = (
    ("rent_deposit", "임차보증금"),
    ("monthly_rent", "월세"),
)
VARIABLE_COST_ACCOUNTS = {"운반비"}


def allowed_file(filename: str) -> bool:
    return Path(filename).suffix.lower() in ALLOWED_EXTENSIONS


def column_name(index: int) -> str:
    index += 1
    letters: list[str] = []
    while index:
        index, remainder = divmod(index - 1, 26)
        letters.append(chr(65 + remainder))
    return "".join(reversed(letters))


def format_number(value: float) -> str:
    if not math.isfinite(value):
        return str(value)
    if value.is_integer():
        return f"{int(value):,}"
    return f"{value:,.10f}".rstrip("0").rstrip(".")


def format_date_value(value: datetime | date) -> str:
    if isinstance(value, datetime):
        if value.time().replace(microsecond=0).isoformat() == "00:00:00":
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return value.strftime("%Y-%m-%d")


def compact_text(value: str) -> str:
    return WHITESPACE_RE.sub("", value)


def split_account_text(value: str | None) -> tuple[str, str]:
    compacted = compact_text(str(value or ""))
    match = ACCOUNT_CODE_RE.match(compacted)
    if not match:
        return "", compacted
    return match.group("code"), compacted[match.end() :]


def normalize_account_text(value: str | None) -> str:
    return split_account_text(value)[1]


def parse_number_text(value: str) -> float | None:
    text = value.replace(",", "").strip()
    if not text:
        return None
    is_parenthesized_negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()")
    try:
        number = float(text)
        return -number if is_parenthesized_negative else number
    except ValueError:
        return None


def display_number(value: float | None) -> str:
    if value is None:
        return ""
    return format_number(value)


def display_whole_number(value: float | None) -> str:
    if value is None or not math.isfinite(value):
        return ""
    return f"{int(value):,}"


def display_percent(value: float | None) -> str:
    if value is None or not math.isfinite(value):
        return ""
    return f"{value:.2f}%"


def parse_percent_text(value: str) -> float | None:
    return parse_number_text(value.replace("%", ""))


def clean_contact_text(value: str | None, max_length: int = 200) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()[:max_length]


def consultation_form_from_request() -> dict[str, Any]:
    return {
        "company": clean_contact_text(request.form.get("company"), 120),
        "contact_name": clean_contact_text(request.form.get("contact_name"), 80),
        "phone": clean_contact_text(request.form.get("phone"), 40),
        "email": clean_contact_text(request.form.get("email"), 120),
        "message": clean_contact_text(request.form.get("message"), 1000),
        "privacy_consent": request.form.get("privacy_consent") == "on",
        "financial_consent": request.form.get("financial_consent") == "on",
    }


def validate_consultation_form(form: dict[str, Any]) -> list[str]:
    # Temporary marketing-flow setting: allow users to unlock the detailed
    # result without completing the contact form while the gate copy is tested.
    return []

    errors: list[str] = []
    if not form["company"]:
        errors.append("회사명을 입력해 주세요.")
    if not form["contact_name"]:
        errors.append("담당자명을 입력해 주세요.")
    if not form["phone"]:
        errors.append("연락처를 입력해 주세요.")
    if not form["email"]:
        errors.append("이메일을 입력해 주세요.")
    elif not EMAIL_RE.match(form["email"]):
        errors.append("이메일 형식을 확인해 주세요.")
    if not form["privacy_consent"]:
        errors.append("개인정보 수집 및 이용에 동의해 주세요.")
    if not form["financial_consent"]:
        errors.append("재무자료 검토 및 상담 목적 이용에 동의해 주세요.")
    return errors


def append_consultation_log(case: dict[str, Any], form: dict[str, Any], result: dict[str, Any]) -> None:
    LEADS_DIR.mkdir(parents=True, exist_ok=True)
    summary = result.get("summary", {})
    payload = {
        "submitted_at": now_iso(),
        "case_id": case["case_id"],
        "uploaded_company_name": case.get("company_name", ""),
        "contact": {
            "company": form["company"],
            "contact_name": form["contact_name"],
            "phone": form["phone"],
            "email": form["email"],
            "message": form["message"],
        },
        "consents": {
            "privacy_consent": form["privacy_consent"],
            "financial_consent": form["financial_consent"],
        },
        "summary": {
            "liquidation_value": summary.get("liquidation_value", 0),
            "going_concern_value": summary.get("going_concern_value", 0),
            "value_difference": summary.get("value_difference", 0),
            "total_debt": summary.get("total_debt", 0),
        },
    }
    with CONSULTATION_LOG_PATH.open("a", encoding="utf-8") as file:
        file.write(json.dumps(payload, ensure_ascii=False) + "\n")


def normalize_email(email: str | None) -> str:
    return str(email or "").strip().lower()


def current_user_id() -> str | None:
    user_id = session.get("user_id")
    return str(user_id) if user_id in USERS else None


def current_user() -> dict[str, Any] | None:
    user_id = current_user_id()
    return USERS.get(user_id) if user_id else None


def find_user_by_email(email: str) -> dict[str, Any] | None:
    normalized = normalize_email(email)
    return next((user for user in USERS.values() if user.get("email") == normalized), None)


def email_serializer() -> URLSafeTimedSerializer:
    return URLSafeTimedSerializer(app.secret_key)


def email_verification_token(user: dict[str, Any]) -> str:
    return email_serializer().dumps(
        {
            "user_id": user.get("user_id"),
            "email": normalize_email(user.get("email")),
        },
        salt=EMAIL_VERIFICATION_SALT,
    )


def load_email_verification_token(token: str) -> dict[str, Any]:
    return email_serializer().loads(
        token,
        salt=EMAIL_VERIFICATION_SALT,
        max_age=EMAIL_VERIFICATION_MAX_AGE_SECONDS,
    )


def app_external_url(endpoint: str, **values: Any) -> str:
    public_url = os.environ.get("RESTART_PUBLIC_URL", "").strip().rstrip("/")
    path = url_for(endpoint, **values)
    if public_url:
        return f"{public_url}{path}"
    return url_for(endpoint, _external=True, **values)


def send_email_verification(user: dict[str, Any]) -> tuple[bool, str]:
    source_email = os.environ.get("SES_SOURCE_EMAIL") or os.environ.get("RESTART_EMAIL_FROM")
    if not source_email:
        return False, "SES_SOURCE_EMAIL 또는 RESTART_EMAIL_FROM 환경변수를 설정해 주세요."

    try:
        import boto3
    except ImportError:
        return False, "boto3 패키지가 설치되어 있지 않습니다."

    token = email_verification_token(user)
    verification_url = app_external_url("verify_email", token=token)
    region_name = os.environ.get("SES_REGION", "ap-northeast-2")
    client = boto3.client("ses", region_name=region_name)
    subject = "[Re-Start] 이메일 인증을 완료해 주세요"
    text_body = (
        "Re-Start 이메일 인증 안내\n\n"
        f"아래 링크를 열어 이메일 인증을 완료해 주세요.\n{verification_url}\n\n"
        "이 링크는 7일 동안 유효합니다."
    )
    html_body = (
        "<p>Re-Start 이메일 인증 안내입니다.</p>"
        "<p>아래 링크를 열어 이메일 인증을 완료해 주세요.</p>"
        f'<p><a href="{verification_url}">이메일 인증하기</a></p>'
        "<p>이 링크는 7일 동안 유효합니다.</p>"
    )

    try:
        client.send_email(
            Source=source_email,
            Destination={"ToAddresses": [normalize_email(user.get("email"))]},
            Message={
                "Subject": {"Data": subject, "Charset": "UTF-8"},
                "Body": {
                    "Text": {"Data": text_body, "Charset": "UTF-8"},
                    "Html": {"Data": html_body, "Charset": "UTF-8"},
                },
            },
        )
    except Exception as exc:
        return False, str(exc)

    user["email_verification_sent_at"] = now_iso()
    return True, ""


def remember_session_case(case_id: str) -> None:
    case_ids = list(session.get("case_ids", []))
    if case_id not in case_ids:
        case_ids.append(case_id)
        session["case_ids"] = case_ids


def attach_session_cases_to_user(user_id: str) -> None:
    for case_id in session.get("case_ids", []):
        case = CASE_STORE.get(case_id)
        if not case:
            continue
        if not case.get("user_id"):
            case["user_id"] = user_id
            save_case(case)


def login_user(user: dict[str, Any]) -> None:
    session["user_id"] = user["user_id"]
    attach_session_cases_to_user(user["user_id"])


def case_is_accessible(case: dict[str, Any]) -> bool:
    if current_admin():
        return True
    owner_id = case.get("user_id")
    if not owner_id:
        return case.get("case_id") in session.get("case_ids", [])
    return owner_id == current_user_id()


def get_accessible_case(case_id: str) -> dict[str, Any] | None:
    case = CASE_STORE.get(case_id)
    if not case or not case_is_accessible(case):
        return None
    return case


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not current_user():
            return redirect(url_for("login", next=request.path))
        return view(*args, **kwargs)

    return wrapped


def admin_username() -> str:
    return os.environ.get("RESTART_ADMIN_USERNAME", "admin").strip() or "admin"


def admin_password_is_configured() -> bool:
    return bool(
        os.environ.get("RESTART_ADMIN_PASSWORD_HASH", "").strip()
        or os.environ.get("RESTART_ADMIN_PASSWORD", "")
    )


def admin_password_matches(password: str) -> bool:
    password_hash = os.environ.get("RESTART_ADMIN_PASSWORD_HASH", "").strip()
    if password_hash:
        return check_password_hash(password_hash, password)

    plain_password = os.environ.get("RESTART_ADMIN_PASSWORD", "")
    return bool(plain_password) and hmac.compare_digest(plain_password, password)


def current_admin() -> str | None:
    admin = session.get("admin_user")
    if not admin:
        return None
    return str(admin)


def admin_login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not current_admin():
            return redirect(url_for("admin_login", next=request.path))
        return view(*args, **kwargs)

    return wrapped


@app.context_processor
def inject_auth_context():
    return {
        "current_user": current_user(),
        "current_admin": current_admin(),
        "format_kst_datetime": format_kst_datetime,
        "format_number": display_whole_number,
    }


def is_editable_financial_account(account: str) -> bool:
    compacted = normalize_account_text(account)
    if not compacted or compacted in {"자산", *ASSET_TOTAL_ACCOUNTS}:
        return False
    return NON_EDITABLE_ACCOUNT_RE.match(compacted) is None


def is_deductible_asset_account(account: str) -> bool:
    compacted = normalize_account_text(account)
    return any(keyword in compacted for keyword in DEDUCTIBLE_ASSET_ACCOUNTS)


def is_numbered_section_account(account: str) -> bool:
    return NON_EDITABLE_ACCOUNT_RE.match(normalize_account_text(account)) is not None


def subtract_from_financial_row(row: dict[str, Any], deduction: float) -> None:
    amount_number = row.get("amount_number")
    if amount_number is not None:
        row["amount_number"] = amount_number - deduction
        row["amount"] = display_number(row["amount_number"])

    if row.get("account") in PREPAID_ACCOUNTS:
        row["audit_value"] = display_number(0.0)
        row["liquidation_value"] = display_number(0.0)
        return

    for key in ("audit_value", "liquidation_value"):
        number = parse_number_text(row.get(key, ""))
        if number is not None:
            row[key] = display_number(number - deduction)


def current_amount_from_balance_row(row: dict[str, Any]) -> str:
    current_cells = [
        cell
        for cell in row["cells"]
        if cell.get("source_col") in {1, 2} and compact_text(cell["text"])
    ]
    for cell in current_cells:
        if parse_number_text(cell["text"]) is not None:
            return cell["text"]
    return current_cells[0]["text"] if current_cells else ""


def extract_financial_rows(sheet: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    last_editable_asset_row: dict[str, Any] | None = None
    in_intangible_asset_section = False

    for row in sheet["rows"]:
        subject_cell = next(
            (cell for cell in row["cells"] if cell.get("source_col") == 0),
            None,
        )
        account_code, account = split_account_text(subject_cell["text"] if subject_cell else "")
        if not account or account in {"과목", "금액"}:
            continue
        if rows and account in LIABILITY_SECTION_START_ACCOUNTS:
            break
        if is_numbered_section_account(account):
            in_intangible_asset_section = INTANGIBLE_ASSET_ACCOUNT in account

        amount_text = current_amount_from_balance_row(row)
        amount_number = parse_number_text(amount_text)
        if is_deductible_asset_account(account):
            if last_editable_asset_row is not None and amount_number is not None:
                subtract_from_financial_row(last_editable_asset_row, abs(amount_number))
            continue

        editable = is_editable_financial_account(account)

        if editable and account in PREPAID_ACCOUNTS:
            audit_value = 0.0
        elif editable and in_intangible_asset_section:
            audit_value = 0.0
        elif editable:
            audit_value = amount_number
        else:
            audit_value = None

        liquidation_value = audit_value if editable else None

        rows.append(
            {
                "row": row["index"],
                "source_row": row.get("source_index", row["index"]),
                "account_code": account_code,
                "account": account,
                "amount": display_number(amount_number),
                "amount_number": amount_number,
                "audit_value": display_number(audit_value),
                "liquidation_value": display_number(liquidation_value),
                "is_editable": editable,
            }
        )
        if editable:
            last_editable_asset_row = rows[-1]
        if account in ASSET_TOTAL_ACCOUNTS:
            break

    return rows


def income_roman_stage(account: str) -> int | None:
    compacted = normalize_account_text(account)
    unicode_stages = (
        ("Ⅰ", 1),
        ("Ⅱ", 2),
        ("Ⅲ", 3),
        ("Ⅳ", 4),
        ("Ⅴ", 5),
        ("Ⅵ", 6),
        ("Ⅶ", 7),
        ("Ⅷ", 8),
        ("Ⅸ", 9),
        ("Ⅹ", 10),
    )
    for prefix, stage in unicode_stages:
        if compacted.startswith(prefix):
            return stage

    upper = compacted.upper()
    ascii_stages = (
        ("VIII", 8),
        ("VII", 7),
        ("VI", 6),
        ("IV", 4),
        ("IX", 9),
        ("III", 3),
        ("II", 2),
        ("V", 5),
        ("X", 10),
        ("I", 1),
    )
    for prefix, stage in ascii_stages:
        if upper.startswith(f"{prefix}.") or upper.startswith(prefix):
            return stage

    return None


def sum_amount_group(row: dict[str, Any], source_cols: tuple[int, ...]) -> tuple[str, float | None]:
    total = 0.0
    found_number = False
    fallback_text = ""

    for cell in row["cells"]:
        if not cell_overlaps_cols(cell, source_cols):
            continue

        text = compact_text(cell["text"])
        if not text:
            continue

        number = parse_number_text(text)
        if number is None:
            fallback_text = fallback_text or text
            continue

        total += number
        found_number = True

    if found_number:
        return display_number(total), total
    return fallback_text, None


def remove_income_rows_between_sales_roman(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    positions: dict[int, int] = {}
    for index, row in enumerate(rows):
        stage = income_roman_stage(row["account"])
        if stage in {1, 2, 3} and stage not in positions:
            positions[stage] = index

    remove_indexes: set[int] = set()
    for previous_stage, current_stage in ((1, 2), (2, 3)):
        previous_index = positions.get(previous_stage)
        current_index = positions.get(current_stage)
        if previous_index is None or current_index is None:
            continue
        if previous_index < current_index - 1:
            remove_indexes.update(range(previous_index + 1, current_index))

    return [row for index, row in enumerate(rows) if index not in remove_indexes]


def is_editable_income_row(account: str, section: str) -> bool:
    compacted = normalize_account_text(account)
    if not compacted:
        return False
    if section == "sales":
        return "매출액" in compacted or "매출원가" in compacted
    return NON_EDITABLE_ACCOUNT_RE.match(compacted) is None


def income_percentage(numerator: float | None, denominator: float | None) -> float | None:
    if numerator is None or denominator in (None, 0):
        return None
    return (numerator / denominator) * 100


def extract_income_rows(sheet: dict[str, Any]) -> list[dict[str, Any]]:
    raw_rows: list[dict[str, Any]] = []

    for row in sheet["rows"]:
        subject_cell = next(
            (cell for cell in row["cells"] if cell.get("source_col") == 0),
            None,
        )
        account_code, account = split_account_text(subject_cell["text"] if subject_cell else "")
        if not account or account in {"과목", "금액"}:
            continue
        if "영업외" in account:
            break

        first_display, first_number = sum_amount_group(row, (1, 2))
        second_display, second_number = sum_amount_group(row, (3, 4))
        average_number = ((first_number or 0) + (second_number or 0)) / 2

        raw_rows.append(
            {
                "row": len(raw_rows) + 1,
                "source_row": row.get("source_index", row["index"]),
                "account_code": account_code,
                "account": account,
                "y_minus_1": first_display,
                "y_minus_1_number": first_number,
                "y": second_display,
                "y_number": second_number,
                "average": display_number(average_number),
                "average_number": average_number,
            }
        )

    filtered_rows = remove_income_rows_between_sales_roman(raw_rows)
    sales_average = next(
        (
            row["average_number"]
            for row in filtered_rows
            if "매출액" in row["account"]
        ),
        None,
    )

    income_rows: list[dict[str, Any]] = []
    section = "sales"
    for row in filtered_rows:
        stage = income_roman_stage(row["account"])
        if stage is not None and stage >= 4:
            section = "expense"

        account = row["account"]
        editable = is_editable_income_row(account, section)
        metric_display = ""
        monthly_average_display = ""
        cost_type = ""
        value_kind = "amount"
        final_value = ""

        if section == "sales":
            value_kind = "percent"
            if "매출액" in account:
                metric_display = display_percent(
                    income_percentage(
                        (row["y_number"] or 0) - (row["y_minus_1_number"] or 0),
                        row["y_minus_1_number"],
                    )
                )
            elif "매출원가" in account:
                metric_display = display_percent(
                    income_percentage(row["average_number"], sales_average)
                )
            final_value = metric_display if editable else ""
        elif editable:
            cost_type = (
                "variable"
                if any(variable in account for variable in VARIABLE_COST_ACCOUNTS)
                else "fixed"
            )
            if cost_type == "variable":
                value_kind = "percent"
                monthly_average_display = display_percent(
                    income_percentage(row["average_number"], sales_average)
                )
            else:
                monthly_average_display = display_whole_number(row["average_number"] / 12)
            final_value = monthly_average_display
        else:
            monthly_average_display = display_whole_number(row["average_number"] / 12)

        income_rows.append(
            {
                **row,
                "row": len(income_rows) + 1,
                "section": section,
                "metric_display": metric_display,
                "monthly_average_display": monthly_average_display,
                "sales_average_number": sales_average,
                "cost_type": cost_type,
                "final_value": final_value,
                "value_kind": value_kind,
                "is_editable": editable,
            }
        )

    return income_rows


def company_name_from_workbook(workbook: dict[str, Any]) -> str:
    for file in workbook["files"]:
        name = Path(file["name"]).stem
        name = re.sub(r"^\d{4}년\d{1,2}월", "", name)
        name = re.sub(r"(재무상태표|손익계산서)", "", name)
        name = name.strip("_ -")
        if name:
            return name
    return "회사"


def create_case(workbook: dict[str, Any]) -> dict[str, Any]:
    case_id = uuid.uuid4().hex[:12]
    created_at = now_iso()
    balance_sheet = next(
        (sheet for sheet in workbook["sheets"] if sheet.get("document_label") == "재무제표"),
        workbook["sheets"][0],
    )
    income_sheet = next(
        (sheet for sheet in workbook["sheets"] if sheet.get("document_label") == "손익계산서"),
        workbook["sheets"][-1],
    )
    case = {
        "case_id": case_id,
        "company_name": company_name_from_workbook(workbook),
        "scenario_name": "",
        "source_case_id": None,
        "user_id": current_user_id(),
        "created_at": created_at,
        "updated_at": created_at,
        "workbook": workbook,
        "financial_rows": extract_financial_rows(balance_sheet),
        "financial_saved": False,
        "debt_rows": build_default_debt_rows(),
        "debt_saved": False,
        "income_rows": extract_income_rows(income_sheet),
        "income_saved": False,
        "collateral_rows": build_default_collateral_rows(),
        "rent_rows": build_default_rent_rows(),
        "collateral_saved": False,
    }
    CASE_STORE[case_id] = case
    remember_session_case(case_id)
    save_case(case)
    return case


def build_default_debt_rows() -> list[dict[str, Any]]:
    return [
        {
            "field": field,
            "category": category,
            "debt_amount": "",
            "collateral_type": "",
            "audit_value": "",
        }
        for field, category in DEBT_DEFAULT_ROWS
    ]


def build_default_collateral_rows() -> list[dict[str, Any]]:
    return [
        {
            "field": field,
            "category": category,
            "audit_value": "",
            "liquidation_value": "",
        }
        for field, category in COLLATERAL_DEFAULT_ROWS
    ]


def build_default_rent_rows() -> list[dict[str, Any]]:
    return [
        {
            "field": field,
            "category": category,
            "amount": "",
        }
        for field, category in RENT_DEFAULT_ROWS
    ]


def clean_submitted_number(value: str) -> tuple[str, float | None]:
    number = parse_number_text(value)
    return display_number(number), number


def clean_submitted_income_value(value: str, value_kind: str) -> tuple[str, float | None]:
    if value_kind == "percent":
        number = parse_percent_text(value)
        return display_percent(number), number
    return clean_submitted_number(value)


def cell_end_col(cell: dict[str, Any]) -> int:
    return cell["source_col"] + max(1, cell.get("colspan", 1)) - 1


def cell_overlaps_cols(cell: dict[str, Any], cols: tuple[int, ...]) -> bool:
    if not cols:
        return False
    return cell["source_col"] <= max(cols) and cell_end_col(cell) >= min(cols)


def cell_overlaps_group(cell: dict[str, Any], group: tuple[int, int]) -> bool:
    return cell["source_col"] <= group[1] and cell_end_col(cell) >= group[0]


def normalize_label_cells(row: dict[str, Any]) -> None:
    for cell in row["cells"]:
        compacted = compact_text(cell["text"])
        if compacted in {"과목", "금액"}:
            cell["text"] = compacted


def normalize_subject_cell(row: dict[str, Any], layout: dict[str, Any]) -> None:
    for cell in row["cells"]:
        if cell_overlaps_cols(cell, layout["subject_cols"]):
            cell["text"] = normalize_account_text(cell["text"])


def subject_text(row: dict[str, Any], layout: dict[str, Any]) -> str:
    texts: list[str] = []
    for cell in row["cells"]:
        if cell_overlaps_cols(cell, layout["subject_cols"]):
            text = normalize_account_text(cell["text"])
            if text:
                texts.append(text)
    return "".join(texts)


def has_amount_label(row: dict[str, Any]) -> bool:
    return any(compact_text(cell["text"]) == "금액" for cell in row["cells"])


def is_metadata_subject(subject: str) -> bool:
    compacted = compact_text(subject)
    if not compacted:
        return False
    if compacted in {"재무상태표", "손익계산서"}:
        return True
    if compacted.startswith(("회사명", "단위", "(단위")):
        return True
    return compacted.startswith("제") and any(
        marker in compacted for marker in ("현재", "부터", "까지")
    )


def detect_table_layout(rows: list[dict[str, Any]], col_count: int) -> dict[str, Any]:
    header_row_index = 0
    subject_cell: dict[str, Any] | None = None

    for row_index, row in enumerate(rows):
        for cell in row["cells"]:
            if compact_text(cell["text"]) == "과목":
                header_row_index = row_index
                subject_cell = cell
                break
        if subject_cell is not None:
            break

    if subject_cell is None:
        return {
            "header_row_index": 0,
            "amount_row_index": 1,
            "subject_cols": (0,),
            "amount_groups": ((1, 1), (2, 2)),
        }

    subject_start = subject_cell["source_col"]
    subject_end = cell_end_col(subject_cell)
    amount_row_index = min(header_row_index + 1, len(rows) - 1)
    amount_groups: list[tuple[int, int]] = []

    for row_index in range(header_row_index + 1, min(header_row_index + 4, len(rows))):
        candidates = [
            cell
            for cell in rows[row_index]["cells"]
            if compact_text(cell["text"]) == "금액"
        ]
        if candidates:
            amount_row_index = row_index
            amount_groups = [
                (cell["source_col"], cell_end_col(cell)) for cell in candidates
            ]
            break

    if not amount_groups:
        amount_start = min(subject_end + 1, max(0, col_count - 1))
        amount_groups = ((amount_start, amount_start),)

    return {
        "header_row_index": header_row_index,
        "amount_row_index": amount_row_index,
        "subject_cols": tuple(range(subject_start, subject_end + 1)),
        "amount_groups": tuple(amount_groups),
    }


def clean_sheet_rows(
    rows: list[dict[str, Any]],
    layout: dict[str, Any],
) -> tuple[list[dict[str, Any]], int]:
    cleaned_rows: list[dict[str, Any]] = []
    removed_count = 0
    table_started = False
    keep_first_amount_header = False

    for row in rows:
        subject = subject_text(row, layout)

        if not table_started:
            if subject == "과목":
                table_started = True
                keep_first_amount_header = True
                normalize_subject_cell(row, layout)
                normalize_label_cells(row)
                cleaned_rows.append(row)
            else:
                removed_count += 1
            continue

        if subject == "과목":
            removed_count += 1
            keep_first_amount_header = False
            continue

        if is_metadata_subject(subject):
            removed_count += 1
            keep_first_amount_header = False
            continue

        if keep_first_amount_header and has_amount_label(row):
            normalize_label_cells(row)
            cleaned_rows.append(row)
            keep_first_amount_header = False
            continue

        keep_first_amount_header = False

        if not subject:
            removed_count += 1
            continue

        if subject.startswith("회사명"):
            removed_count += 1
            continue

        normalize_subject_cell(row, layout)
        cleaned_rows.append(row)

    for display_index, row in enumerate(cleaned_rows, start=1):
        row["index"] = display_index

    return cleaned_rows, removed_count


def cell_by_source_col(row: dict[str, Any], source_col: int) -> dict[str, Any] | None:
    for cell in row["cells"]:
        if cell.get("source_col") == source_col:
            return cell
    return None


def is_empty_cell(cell: dict[str, Any] | None) -> bool:
    return cell is None or compact_text(cell["text"]) == ""


def empty_cell(source_col: int) -> dict[str, Any]:
    return {
        "text": "",
        "kind": "empty",
        "rowspan": 1,
        "colspan": 1,
        "source_col": source_col,
    }


def display_cell(cell: dict[str, Any] | None, source_col: int) -> dict[str, Any]:
    if cell is None:
        return empty_cell(source_col)
    return {
        **cell,
        "rowspan": 1,
        "colspan": 1,
        "source_col": source_col,
    }


def make_text_cell(
    text: str,
    source_col: int,
    colspan: int = 1,
    rowspan: int = 1,
) -> dict[str, Any]:
    return {
        "text": text,
        "kind": "text" if text else "empty",
        "rowspan": rowspan,
        "colspan": colspan,
        "source_col": source_col,
    }


def first_cell_text_for_group(row: dict[str, Any], group: tuple[int, int]) -> str:
    for cell in row["cells"]:
        if cell_overlaps_group(cell, group):
            text = compact_text(cell["text"])
            if text:
                return text
    return ""


def subject_display_cell(row: dict[str, Any], layout: dict[str, Any]) -> dict[str, Any]:
    text = subject_text(row, layout)
    source = next(
        (cell for cell in row["cells"] if cell_overlaps_cols(cell, layout["subject_cols"])),
        None,
    )
    return {
        "text": text,
        "kind": "text" if text else "empty",
        "rowspan": source.get("rowspan", 1) if source else 1,
        "colspan": 1,
        "source_col": 0,
    }


def physical_cells_for_group(
    row: dict[str, Any],
    group: tuple[int, int],
    display_start_col: int,
) -> list[dict[str, Any]]:
    cells: list[dict[str, Any]] = []
    for offset, source_col in enumerate(range(group[0], group[1] + 1)):
        cells.append(display_cell(cell_by_source_col(row, source_col), display_start_col + offset))
    return cells


def normalize_income_statement_sheet(sheet: dict[str, Any]) -> None:
    layout = sheet.get("layout")
    if not layout or len(layout["amount_groups"]) < 2 or len(sheet["rows"]) < 2:
        return

    compacted_rows: list[dict[str, Any]] = []
    amount_groups = layout["amount_groups"][:2]
    group_widths = [2, 2]

    for row_index, row in enumerate(sheet["rows"]):
        if row_index == 0:
            compacted_rows.append(
                {
                    **row,
                    "cells": [
                        subject_display_cell(row, layout),
                        make_text_cell(
                            first_cell_text_for_group(row, amount_groups[0]),
                            1,
                            colspan=2,
                        ),
                        make_text_cell(
                            first_cell_text_for_group(row, amount_groups[1]),
                            3,
                            colspan=2,
                        ),
                    ],
                }
            )
            continue

        if row_index == 1:
            compacted_rows.append(
                {
                    **row,
                    "cells": [
                        make_text_cell("금액", 1, colspan=2),
                        make_text_cell("금액", 3, colspan=2),
                    ],
                }
            )
            continue

        cells = [subject_display_cell(row, layout)]
        display_col = 1
        for group, width in zip(amount_groups, group_widths):
            group_cells = physical_cells_for_group(row, group, display_col)
            if len(group_cells) < width:
                group_cells.extend(empty_cell(display_col + i) for i in range(len(group_cells), width))
            cells.extend(group_cells[:width])
            display_col += width

        compacted_rows.append(
            {
                **row,
                "cells": cells,
            }
        )

    sheet["rows"] = compacted_rows
    sheet["col_count"] = 5
    sheet["columns"] = ["A", "B", "C", "D", "E"]


def normalize_balance_sheet(sheet: dict[str, Any]) -> None:
    layout = sheet.get("layout")
    if not layout or len(layout["amount_groups"]) < 2 or len(sheet["rows"]) < 2:
        return

    amount_groups = layout["amount_groups"][:2]
    group_widths = [group[1] - group[0] + 1 for group in amount_groups]
    compacted_rows: list[dict[str, Any]] = []

    for row_index, row in enumerate(sheet["rows"]):
        cells: list[dict[str, Any]] = []

        if row_index == 0:
            cells.append(subject_display_cell(row, layout))
            display_col = 1
            for group, width in zip(amount_groups, group_widths):
                cells.append(
                    make_text_cell(
                        first_cell_text_for_group(row, group),
                        display_col,
                        colspan=width,
                    )
                )
                display_col += width
        elif row_index == 1:
            display_col = 1
            for width in group_widths:
                cells.append(make_text_cell("금액", display_col, colspan=width))
                display_col += width
        else:
            cells.append(subject_display_cell(row, layout))
            display_col = 1
            for group in amount_groups:
                group_cells = physical_cells_for_group(row, group, display_col)
                cells.extend(group_cells)
                display_col += len(group_cells)

        compacted_rows.append({**row, "cells": cells})

    col_count = 1 + sum(group_widths)
    sheet["rows"] = compacted_rows
    sheet["col_count"] = col_count
    sheet["columns"] = [column_name(index) for index in range(col_count)]


def normalize_financial_sheet(sheet: dict[str, Any], document_label: str) -> None:
    if document_label == "손익계산서":
        normalize_income_statement_sheet(sheet)
    else:
        normalize_balance_sheet(sheet)


def build_merge_maps(
    merged_ranges: list[tuple[int, int, int, int]],
) -> tuple[dict[tuple[int, int], tuple[int, int]], set[tuple[int, int]]]:
    top_left: dict[tuple[int, int], tuple[int, int]] = {}
    covered: set[tuple[int, int]] = set()

    for row_start, row_end, col_start, col_end in merged_ranges:
        rowspan = max(1, row_end - row_start)
        colspan = max(1, col_end - col_start)
        top_left[(row_start, col_start)] = (rowspan, colspan)

        for row in range(row_start, row_end):
            for col in range(col_start, col_end):
                if (row, col) != (row_start, col_start):
                    covered.add((row, col))

    return top_left, covered


def format_xls_cell(book: xlrd.book.Book, cell: xlrd.sheet.Cell) -> tuple[str, str]:
    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return "", "empty"
    if cell.ctype == xlrd.XL_CELL_TEXT:
        return str(cell.value).strip(), "text"
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        return format_number(float(cell.value)), "number"
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            value = xlrd.xldate.xldate_as_datetime(cell.value, book.datemode)
            return format_date_value(value), "date"
        except (OverflowError, ValueError):
            return str(cell.value), "text"
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return ("TRUE" if cell.value else "FALSE"), "boolean"
    if cell.ctype == xlrd.XL_CELL_ERROR:
        return "#ERROR", "error"
    return str(cell.value), "text"


def parse_xls(
    file_bytes: bytes,
    display_sheet_name: str | None = None,
    only_first_sheet: bool = False,
) -> dict[str, Any]:
    try:
        book = xlrd.open_workbook(file_contents=file_bytes, formatting_info=True)
    except NotImplementedError:
        book = xlrd.open_workbook(file_contents=file_bytes)

    sheets = []
    source_sheets = [book.sheet_by_index(0)] if only_first_sheet else book.sheets()
    for sheet in source_sheets:
        merged_ranges = list(sheet.merged_cells)
        merge_top_left, merge_covered = build_merge_maps(merged_ranges)

        rows = []
        for row_index in range(sheet.nrows):
            rendered_cells = []
            for col_index in range(sheet.ncols):
                if (row_index, col_index) in merge_covered:
                    continue

                text, kind = format_xls_cell(book, sheet.cell(row_index, col_index))
                rowspan, colspan = merge_top_left.get((row_index, col_index), (1, 1))
                rendered_cells.append(
                    {
                        "text": text,
                        "kind": kind,
                        "rowspan": rowspan,
                        "colspan": colspan,
                        "source_col": col_index,
                    }
                )

            rows.append(
                {
                    "index": row_index + 1,
                    "source_index": row_index + 1,
                    "cells": rendered_cells,
                }
            )

        layout = detect_table_layout(rows, sheet.ncols)
        rows, removed_count = clean_sheet_rows(rows, layout)

        sheets.append(
            {
                "name": display_sheet_name or sheet.name,
                "source_sheet_name": sheet.name,
                "row_count": len(rows),
                "original_row_count": sheet.nrows,
                "removed_row_count": removed_count,
                "col_count": sheet.ncols,
                "columns": [column_name(index) for index in range(sheet.ncols)],
                "layout": layout,
                "rows": rows,
            }
        )

    return {"sheets": sheets}


def format_xlsx_value(value: Any) -> tuple[str, str]:
    if value is None:
        return "", "empty"
    if isinstance(value, bool):
        return ("TRUE" if value else "FALSE"), "boolean"
    if isinstance(value, (datetime, date)):
        return format_date_value(value), "date"
    if isinstance(value, int):
        return f"{value:,}", "number"
    if isinstance(value, float):
        return format_number(value), "number"
    return str(value).strip(), "text"


def parse_xlsx(
    file_bytes: bytes,
    display_sheet_name: str | None = None,
    only_first_sheet: bool = False,
) -> dict[str, Any]:
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=False, data_only=True)
    sheets = []

    source_worksheets = [workbook.worksheets[0]] if only_first_sheet else workbook.worksheets
    for worksheet in source_worksheets:
        max_row = worksheet.max_row or 1
        max_col = worksheet.max_column or 1
        merged_ranges = [
            (
                merged.min_row - 1,
                merged.max_row,
                merged.min_col - 1,
                merged.max_col,
            )
            for merged in worksheet.merged_cells.ranges
        ]
        merge_top_left, merge_covered = build_merge_maps(merged_ranges)

        rows = []
        for row_index in range(max_row):
            rendered_cells = []
            for col_index in range(max_col):
                if (row_index, col_index) in merge_covered:
                    continue

                cell = worksheet.cell(row=row_index + 1, column=col_index + 1)
                text, kind = format_xlsx_value(cell.value)
                rowspan, colspan = merge_top_left.get((row_index, col_index), (1, 1))
                rendered_cells.append(
                    {
                        "text": text,
                        "kind": kind,
                        "rowspan": rowspan,
                        "colspan": colspan,
                        "source_col": col_index,
                    }
                )

            rows.append(
                {
                    "index": row_index + 1,
                    "source_index": row_index + 1,
                    "cells": rendered_cells,
                }
            )

        layout = detect_table_layout(rows, max_col)
        rows, removed_count = clean_sheet_rows(rows, layout)

        sheets.append(
            {
                "name": display_sheet_name or worksheet.title,
                "source_sheet_name": worksheet.title,
                "row_count": len(rows),
                "original_row_count": max_row,
                "removed_row_count": removed_count,
                "col_count": max_col,
                "columns": [column_name(index) for index in range(max_col)],
                "layout": layout,
                "rows": rows,
            }
        )

    return {"sheets": sheets}


def parse_workbook(
    filename: str,
    file_bytes: bytes,
    display_sheet_name: str | None = None,
    only_first_sheet: bool = False,
) -> dict[str, Any]:
    extension = Path(filename).suffix.lower()
    if extension == ".xls":
        return parse_xls(file_bytes, display_sheet_name, only_first_sheet)
    if extension == ".xlsx":
        return parse_xlsx(file_bytes, display_sheet_name, only_first_sheet)
    raise ValueError("지원하지 않는 엑셀 형식입니다.")


def build_document_workbook(uploaded_files: Any) -> dict[str, Any]:
    sheets = []
    files = []

    for field_name, document_label in DOCUMENT_FIELDS:
        uploaded_file = uploaded_files.get(field_name)
        if uploaded_file is None or uploaded_file.filename == "":
            raise ValueError(f"{document_label} 파일을 선택해 주세요.")

        display_name = os.path.basename(uploaded_file.filename)
        if not allowed_file(display_name):
            raise ValueError(f"{document_label} 파일은 .xls 또는 .xlsx만 업로드할 수 있습니다.")

        parsed = parse_workbook(
            display_name,
            uploaded_file.read(),
            display_sheet_name=document_label,
            only_first_sheet=True,
        )
        if not parsed["sheets"]:
            raise ValueError(f"{document_label} 파일에서 첫 번째 시트를 찾을 수 없습니다.")

        sheet = parsed["sheets"][0]
        normalize_financial_sheet(sheet, document_label)
        sheet["filename"] = display_name
        sheet["document_label"] = document_label
        sheets.append(sheet)
        files.append({"label": document_label, "name": display_name})

    return {"sheets": sheets, "files": files}


EXCEL_SHEET_INVALID_CHARS_RE = re.compile(r"[:\\/?*\[\]]")


def safe_excel_sheet_title(title: Any, used_titles: set[str]) -> str:
    base = EXCEL_SHEET_INVALID_CHARS_RE.sub(" ", str(title or "Sheet")).strip()
    base = re.sub(r"\s+", " ", base) or "Sheet"
    base = base[:31]
    candidate = base
    index = 2
    while candidate in used_titles:
        suffix = f" {index}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        index += 1
    used_titles.add(candidate)
    return candidate


def safe_download_basename(value: Any, fallback: str = "restart") -> str:
    text = re.sub(r"[\\/:*?\"<>|]+", " ", str(value or fallback)).strip()
    text = re.sub(r"\s+", " ", text) or fallback
    return text[:80].strip()


def extracted_excel_value(cell: dict[str, Any]) -> Any:
    text = str(cell.get("text") or "").strip()
    if not text:
        return None

    kind = cell.get("kind")
    if kind == "number":
        number = parse_number_text(text)
        if number is not None:
            return int(number) if number.is_integer() else number
    if kind == "boolean":
        return text.upper() == "TRUE"
    return text


def style_extracted_worksheet(worksheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="EAF1F0")
    section_fill = PatternFill("solid", fgColor="F5F7F5")
    border_color = "D6DEDB"
    thin_border = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color),
    )

    for row in worksheet.iter_rows():
        non_empty_values = [cell.value for cell in row if cell.value not in (None, "")]
        row_index = row[0].row
        is_header_like = len(non_empty_values) > 1
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        if is_header_like and row_index <= 2:
            for cell in row:
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        elif len(non_empty_values) == 1 and row[0].value not in (None, ""):
            for cell in row:
                cell.fill = section_fill
            row[0].font = Font(bold=True)

    for column_index in range(1, worksheet.max_column + 1):
        max_length = 8
        for cell in worksheet.iter_cols(min_col=column_index, max_col=column_index, values_only=False).__next__():
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value) + 2, 34))
        worksheet.column_dimensions[get_column_letter(column_index)].width = max_length

    for row_index in range(1, worksheet.max_row + 1):
        worksheet.row_dimensions[row_index].height = 24
    worksheet.freeze_panes = "A2"


def excel_display_value(value: Any) -> Any:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return value

    text = str(value).strip()
    if not text:
        return None
    if text.endswith("%"):
        percent = parse_percent_text(text)
        return text if percent is None else percent / 100

    number = parse_number_text(text)
    if number is not None:
        return int(number) if number.is_integer() else number
    return text


def append_table_sheet(
    workbook: Workbook,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    used_titles: set[str],
) -> Any:
    worksheet = workbook.create_sheet(safe_excel_sheet_title(title, used_titles))
    worksheet.append(headers)
    for row in rows:
        worksheet.append([excel_display_value(value) for value in row])

    header_fill = PatternFill("solid", fgColor="EAF1F0")
    border_color = "D6DEDB"
    thin_border = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color),
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                if "율" in str(worksheet.cell(row=1, column=cell.column).value or "") or str(cell.value).endswith("%"):
                    cell.number_format = "0.00%"
                else:
                    cell.number_format = "#,##0"

    for column_index in range(1, worksheet.max_column + 1):
        max_length = 10
        for cell in worksheet.iter_cols(min_col=column_index, max_col=column_index, values_only=False).__next__():
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value) + 2, 42))
        worksheet.column_dimensions[get_column_letter(column_index)].width = max_length

    worksheet.freeze_panes = "A2"
    return worksheet


def append_case_input_sheets(workbook: Workbook, case: dict[str, Any], used_titles: set[str]) -> dict[str, str]:
    asset_sheet = append_table_sheet(
        workbook,
        "자산가치산정",
        ["계정과목", "장부금액", "실사가치", "청산가치", "편집가능"],
        [
            [
                row.get("account", ""),
                row.get("amount", ""),
                row.get("audit_value", ""),
                row.get("liquidation_value", ""),
                "Y" if row.get("is_editable") else "N",
            ]
            for row in case.get("financial_rows", [])
        ],
        used_titles,
    )

    debt_sheet = append_table_sheet(
        workbook,
        "채무입력",
        ["구분", "채무금액"],
        [[row.get("category", ""), row.get("debt_amount", "")] for row in case.get("debt_rows", [])],
        used_titles,
    )

    income_rows = case.get("income_rows", [])
    income_sheet = append_table_sheet(
        workbook,
        "손익추정",
        ["구분", "항목", "전전년도(Y-2)", "전년도(Y-1)", "성장률/원가율 또는 고정/변동", "월평균(기본값)", "최종입력값"],
        [
            [
                "매출" if row.get("section") == "sales" else "판매비와 관리비",
                row.get("account", ""),
                row.get("y_minus_1", ""),
                row.get("y", ""),
                row.get("metric_display") or row.get("cost_type", ""),
                row.get("monthly_average_display", ""),
                row.get("final_value", ""),
            ]
            for row in income_rows
        ],
        used_titles,
    )

    collateral_rows = [
        ["담보등자산", row.get("category", ""), row.get("audit_value", ""), row.get("liquidation_value", "")]
        for row in case.get("collateral_rows", [])
    ]
    rent_rows = [["재임차", row.get("category", ""), row.get("amount", ""), ""] for row in case.get("rent_rows", [])]
    collateral_sheet = append_table_sheet(
        workbook,
        "담보",
        ["구분", "항목", "실사가치/금액", "청산가치"],
        collateral_rows + rent_rows,
        used_titles,
    )
    return {
        "assets": asset_sheet.title,
        "debts": debt_sheet.title,
        "income": income_sheet.title,
        "collateral": collateral_sheet.title,
    }


def append_result_sheets(
    workbook: Workbook,
    result: dict[str, Any],
    used_titles: set[str],
) -> None:
    rows: list[list[Any]] = []
    summary = result.get("summary", {})
    assets = result.get("assets", {})
    sections = assets.get("sections", {})
    current = sections.get("current", {})
    non_current = sections.get("non_current", {})
    enterprise_value = result.get("enterprise_value", {})
    diagnosis = result.get("diagnosis", {})

    rows.extend(
        [
            ["핵심 결과", "청산가치", summary.get("liquidation_value", 0)],
            ["핵심 결과", "계속기업가치", summary.get("going_concern_value", 0)],
            ["핵심 결과", "가치 차이", summary.get("value_difference", 0)],
            ["핵심 결과", "총 채무", summary.get("total_debt", 0)],
            ["자산평가 요약", "유동자산 재무제표상 금액", current.get("statement", 0)],
            ["자산평가 요약", "유동자산 실사가치", current.get("audit", 0)],
            ["자산평가 요약", "유동자산 청산가치", current.get("liquidation", 0)],
            ["자산평가 요약", "비유동자산 재무제표상 금액", non_current.get("statement", 0)],
            ["자산평가 요약", "비유동자산 실사가치", non_current.get("audit", 0)],
            ["자산평가 요약", "비유동자산 청산가치", non_current.get("liquidation", 0)],
            ["가치 비교", "영업활동현금흐름 현재가치", enterprise_value.get("pv_10_years", 0) + enterprise_value.get("terminal_value", 0)],
            ["가치 비교", "비영업용자산 처분가치", enterprise_value.get("non_business_asset_value", 0)],
            ["가치 비교", "계속기업가치", summary.get("going_concern_value", 0)],
            ["가치 비교", "청산가치", summary.get("liquidation_value", 0)],
            ["가치 비교", "차이", summary.get("value_difference", 0)],
            ["진단 결과", "가치 비교", diagnosis.get("value_message", "")],
            ["진단 결과", "변제율 비교", diagnosis.get("repayment_message", "")],
            ["진단 결과", "채권자 동의 가능성", diagnosis.get("consent_message", "")],
            ["진단 결과", "종합 결론", diagnosis.get("overall_message", "")],
        ]
    )
    append_table_sheet(workbook, "결과", ["구분", "항목", "값"], rows, used_titles)

    append_table_sheet(
        workbook,
        "변제율 비교",
        ["구분", "채무금액", "청산 시 변제금액", "청산 시 변제율", "계속기업 변제금액", "계속기업 변제율", "변제금액 차이", "변제율 차이"],
        [
            [
                row.get("label", ""),
                row.get("debt", 0),
                row.get("liquidation_repayment", 0),
                row.get("liquidation_rate", 0),
                row.get("going_repayment", 0),
                row.get("going_rate", 0),
                row.get("repayment_difference", 0),
                row.get("rate_difference", 0),
            ]
            for row in result.get("comparison_rows", [])
        ],
        used_titles,
    )


REVIEW_COLUMNS = ("C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N")
REVIEW_PERIOD_COLUMNS = REVIEW_COLUMNS[:11]


def quote_excel_sheet_name(sheet_name: str) -> str:
    escaped = sheet_name.replace("'", "''")
    return f"'{escaped}'"


def external_cell(sheet_name: str, cell: str) -> str:
    return f"{quote_excel_sheet_name(sheet_name)}!{cell}"


def excel_sum(terms: list[str]) -> str:
    terms = [term for term in terms if term]
    if not terms:
        return "0"
    if len(terms) == 1:
        return terms[0]
    return f"SUM({','.join(terms)})"


def excel_formula_value(value: Any) -> str:
    if isinstance(value, date):
        return f"DATE({value.year},{value.month},{value.day})"
    if value in (None, ""):
        return '""'
    if isinstance(value, (int, float)):
        return repr(float(value))
    escaped = str(value).replace('"', '""')
    return f'"{escaped}"'


def worksheet_formula_sources(case: dict[str, Any], sheet_titles: dict[str, str]) -> dict[str, Any]:
    financial_rows = case.get("financial_rows", [])
    current_rows: list[int] = []
    non_current_rows: list[int] = []
    cash_row = None
    section = None

    for index, row in enumerate(financial_rows, start=2):
        account = str(row.get("account", ""))
        if "비유동자산" in account:
            section = "non_current"
            continue
        if "유동자산" in account:
            section = "current"
            continue
        if section is None or not row.get("is_editable"):
            continue
        if section == "current":
            current_rows.append(index)
        else:
            non_current_rows.append(index)
        if cash_row is None and "현금" in account:
            cash_row = index

    debt_rows = {
        row.get("field"): index
        for index, row in enumerate(case.get("debt_rows", []), start=2)
        if row.get("field")
    }
    income_rows = case.get("income_rows", [])
    income_by_text = {
        text: next((index for index, row in enumerate(income_rows, start=2) if text in str(row.get("account", ""))), None)
        for text in ("매출액", "매출원가", "영업이익")
    }
    expense_rows = [
        (index, row)
        for index, row in enumerate(income_rows, start=2)
        if row.get("section") == "expense" and row.get("is_editable")
    ]
    collateral_rows = {
        row.get("field"): index
        for index, row in enumerate(case.get("collateral_rows", []), start=2)
        if row.get("field")
    }
    rent_start = 2 + len(case.get("collateral_rows", []))
    rent_rows = {
        row.get("field"): index
        for index, row in enumerate(case.get("rent_rows", []), start=rent_start)
        if row.get("field")
    }

    return {
        "sheet_titles": sheet_titles,
        "current_asset_rows": current_rows,
        "non_current_asset_rows": non_current_rows,
        "cash_row": cash_row,
        "debt_rows": debt_rows,
        "income_by_text": income_by_text,
        "expense_rows": expense_rows,
        "collateral_rows": collateral_rows,
        "rent_rows": rent_rows,
    }


def worksheet_review_formulas(
    case: dict[str, Any],
    result: dict[str, Any],
    sheet_titles: dict[str, str],
    row_index_by_number: dict[int, int],
) -> tuple[dict[tuple[int, str], str], dict[tuple[int, str], str]]:
    sources = worksheet_formula_sources(case, sheet_titles)
    formulas: dict[tuple[int, str], str] = {}
    formats: dict[tuple[int, str], str] = {}

    def cell(row_number_value: int, column: str) -> str:
        return f"{column}{row_index_by_number[row_number_value]}"

    def set_formula(row_number_value: int, column: str, formula: str, number_format: str | None = "#,##0") -> None:
        formulas[(row_number_value, column)] = formula if formula.startswith("=") else f"={formula}"
        if number_format:
            formats[(row_number_value, column)] = number_format

    def set_value(row_number_value: int, column: str, value: Any, number_format: str | None = "#,##0") -> None:
        set_formula(row_number_value, column, excel_formula_value(value), number_format)

    def source(sheet_key: str, column: str, row_index: int | None, default: str = "0") -> str:
        if not row_index:
            return default
        return external_cell(sources["sheet_titles"][sheet_key], f"{column}{row_index}")

    def source_sum(sheet_key: str, column: str, row_indexes: list[int]) -> str:
        return excel_sum([source(sheet_key, column, row_index) for row_index in row_indexes])

    def safe_div(numerator: str, denominator: str) -> str:
        return f"IFERROR({numerator}/{denominator},0)"

    current_rows = sources["current_asset_rows"]
    non_current_rows = sources["non_current_asset_rows"]
    for row_number_value, source_rows in ((7, current_rows), (38, non_current_rows)):
        set_formula(row_number_value, "C", source_sum("assets", "B", source_rows))
        set_formula(row_number_value, "D", source_sum("assets", "C", source_rows))
        set_formula(row_number_value, "E", source_sum("assets", "D", source_rows))
    for column in ("C", "D", "E"):
        set_formula(315, column, f"SUM({cell(7, column)},{cell(38, column)})")

    debt_row_map = sources["debt_rows"]
    debt_rows = [
        (71, "secured_debt"),
        (72, "unsecured_financial_debt"),
        (73, "other_unsecured_debt"),
        (74, "related_party_debt"),
        (75, "unpaid_wages"),
        (76, "retirement_benefit"),
        (77, "tax_arrears"),
    ]
    for row_number_value, key in debt_rows:
        set_formula(row_number_value, "C", source("debts", "B", debt_row_map.get(key)))
    set_formula(78, "C", f"SUM({cell(71, 'C')}:{cell(77, 'C')})")

    income_rows = sources["income_by_text"]
    sales_row = income_rows.get("매출액")
    cost_row = income_rows.get("매출원가")
    operating_profit_row = income_rows.get("영업이익")
    for row_number_value, income_row in ((86, sales_row), (87, cost_row)):
        set_formula(row_number_value, "C", source("income", "C", income_row))
        set_formula(row_number_value, "D", source("income", "D", income_row))
        set_formula(row_number_value, "F", source("income", "E", income_row, '""'), None)
        set_formula(row_number_value, "G", source("income", "G", income_row), "0.00%")
    set_formula(129, "C", source("income", "C", operating_profit_row))
    set_formula(129, "D", source("income", "D", operating_profit_row))

    collateral_row_map = sources["collateral_rows"]
    collateral_rows = [
        (136, "collateral_except_machinery"),
        (137, "collateral_machinery"),
        (138, "savings"),
        (139, "insurance"),
        (140, "securities"),
        (141, "other_non_business_assets"),
    ]
    for row_number_value, key in collateral_rows:
        set_formula(row_number_value, "C", source("collateral", "C", collateral_row_map.get(key)))
        set_formula(row_number_value, "D", source("collateral", "D", collateral_row_map.get(key)))
    set_formula(142, "C", f"SUM({cell(136, 'C')}:{cell(141, 'C')})")
    set_formula(142, "D", f"SUM({cell(136, 'D')}:{cell(141, 'D')})")

    rent_row_map = sources["rent_rows"]
    set_formula(146, "C", source("collateral", "C", rent_row_map.get("rent_deposit")))
    set_formula(147, "C", source("collateral", "C", rent_row_map.get("monthly_rent")))

    assumptions = result.get("assumptions")
    dates = result.get("dates", {})
    set_value(161, "C", getattr(assumptions, "cpi_rate", 0.019), "0.00%")
    set_value(162, "C", getattr(assumptions, "bond_yield_3y", 0.02335), "0.00%")
    set_value(163, "C", getattr(assumptions, "repayment_present_value_rate", 0.0542), "0.00%")
    set_value(164, "C", dates.get("prep_days", 0))
    set_value(165, "C", dates.get("total_days", 0))
    set_value(168, "C", dates.get("today", ""), "yyyy-mm-dd")
    set_value(169, "C", dates.get("end", ""), "yyyy-mm-dd")
    set_formula(169, "D", cell(164, "C"))
    set_value(170, "C", dates.get("start", ""), "yyyy-mm-dd")
    set_formula(170, "D", cell(165, "C"))

    set_formula(185, "D", f"SUM({cell(136, 'C')},{cell(137, 'C')})")
    set_formula(185, "E", f"{cell(315, 'E')}-{cell(185, 'D')}")
    set_formula(185, "G", cell(315, "E"))
    set_formula(186, "D", f"-{cell(185, 'D')}*0.05")
    set_formula(186, "E", f"-{cell(185, 'E')}*0.05")
    set_formula(186, "G", f"-{cell(185, 'G')}*0.05")
    for column in ("D", "E", "G"):
        set_formula(187, column, f"{cell(185, column)}+{cell(186, column)}")

    liquidation_debt_sources = {
        189: 75,
        190: 76,
        191: 77,
        192: 71,
        193: 72,
        194: 73,
        195: 74,
    }
    for row_number_value, source_row in liquidation_debt_sources.items():
        set_formula(row_number_value, "C", cell(source_row, "C"))

    set_formula(192, "D", f"MIN({cell(192, 'C')},{cell(187, 'D')})")
    set_formula(189, "D", f"MIN({cell(189, 'C')},{cell(187, 'D')}-{cell(192, 'D')})")
    set_formula(190, "D", f"MIN({cell(190, 'C')},{cell(187, 'D')}-{cell(192, 'D')}-{cell(189, 'D')})")
    set_formula(191, "D", f"MIN({cell(191, 'C')},{cell(187, 'D')}-{cell(192, 'D')}-{cell(189, 'D')}-{cell(190, 'D')})")
    for row_number_value in (193, 194, 195):
        set_value(row_number_value, "D", 0)

    set_formula(189, "E", f"MIN({cell(189, 'C')}-{cell(189, 'D')},{cell(187, 'E')})")
    set_formula(190, "E", f"MIN({cell(190, 'C')}-{cell(190, 'D')},{cell(187, 'E')}-{cell(189, 'E')})")
    set_formula(191, "E", f"MIN({cell(191, 'C')}-{cell(191, 'D')},{cell(187, 'E')}-{cell(189, 'E')}-{cell(190, 'E')})")
    secured_unsecured_denominator = f"SUM({cell(192, 'C')}:{cell(195, 'C')})"
    unsecured_priority_remainder = f"{cell(187, 'E')}-SUM({cell(189, 'E')}:{cell(191, 'E')})"
    for row_number_value in (192, 193, 194, 195):
        set_formula(
            row_number_value,
            "E",
            f"MIN({cell(row_number_value, 'C')}-{cell(row_number_value, 'D')},IFERROR(({unsecured_priority_remainder})*{cell(row_number_value, 'C')}/({secured_unsecured_denominator}),0))",
        )

    residual_pool = f"{cell(187, 'E')}-SUM({cell(189, 'E')}:{cell(195, 'E')})"
    residual_denominator = f"SUM({cell(193, 'C')}:{cell(195, 'C')})"
    for row_number_value in (189, 190, 191, 192):
        set_value(row_number_value, "F", 0)
    for row_number_value in (193, 194, 195):
        set_formula(
            row_number_value,
            "F",
            f"MIN({cell(row_number_value, 'C')}-{cell(row_number_value, 'D')}-{cell(row_number_value, 'E')},IFERROR(({residual_pool})*{cell(row_number_value, 'C')}/({residual_denominator}),0))",
        )
    for row_number_value in range(189, 196):
        set_formula(row_number_value, "G", f"SUM({cell(row_number_value, 'D')}:{cell(row_number_value, 'F')})")
        set_formula(row_number_value, "H", f"{cell(row_number_value, 'C')}-{cell(row_number_value, 'G')}")
        set_formula(row_number_value, "I", safe_div(cell(row_number_value, "G"), cell(row_number_value, "C")), "0.00%")
    for column in ("C", "D", "E", "F", "G", "H"):
        set_formula(196, column, f"SUM({cell(189, column)}:{cell(195, column)})")

    prep_ratio = safe_div(cell(164, "C"), cell(165, "C"))
    set_formula(204, "C", f"IFERROR({cell(86, 'D')}*(1+{cell(86, 'G')})*{cell(164, 'C')}/{cell(165, 'C')},0)")
    set_formula(204, "D", f"IFERROR({cell(204, 'C')}/({prep_ratio})*(1+{cell(86, 'G')}),0)")
    previous_column = "D"
    for column in REVIEW_PERIOD_COLUMNS[2:]:
        set_formula(204, column, f"{cell(204, previous_column)}*(1+{cell(86, 'G')})")
        previous_column = column
    for column in REVIEW_PERIOD_COLUMNS:
        set_formula(205, column, f"{cell(204, column)}*{cell(87, 'G')}")
        set_formula(206, column, f"{cell(204, column)}-{cell(205, column)}")

    expense_rows = sources["expense_rows"]

    def expense_term(row_index: int, row: dict[str, Any], period_index: int, period_column: str) -> str:
        final_value_ref = f"N({source('income', 'G', row_index)})"
        if row.get("cost_type") == "variable":
            return f"{cell(204, period_column)}*{final_value_ref}"
        if period_index == 0:
            return f"{final_value_ref}*12*{prep_ratio}"
        return f"{final_value_ref}*12*(1+{cell(161, 'C')})^{period_index}"

    for period_index, column in enumerate(REVIEW_PERIOD_COLUMNS):
        terms = [
            expense_term(row_index, row, period_index, column)
            for row_index, row in expense_rows
        ]
        rent_term = (
            f"{cell(147, 'C')}*12*{prep_ratio}"
            if period_index == 0
            else f"{cell(147, 'C')}*12*(1+{cell(161, 'C')})^{period_index}"
        )
        terms.append(rent_term)
        set_formula(207, column, excel_sum(terms))
        set_formula(248, column, f"{cell(206, column)}-{cell(207, column)}")
        set_formula(254, column, cell(248, column))
    set_formula(254, "N", f"SUM({cell(254, 'C')}:{cell(254, 'M')})")

    set_formula(258, "C", f"{cell(162, 'C')}+0.065", "0.00%")
    for period_index, column in enumerate(REVIEW_PERIOD_COLUMNS):
        if period_index == 0:
            set_formula(261, column, prep_ratio, "0.0000")
        else:
            previous = REVIEW_PERIOD_COLUMNS[period_index - 1]
            set_formula(261, column, f"{cell(261, previous)}+1", "0.0000")
        set_formula(255, column, f"1/(1+{cell(258, 'C')})^{cell(261, column)}", "0.0000")
        set_formula(262, column, cell(255, column), "0.0000")
        set_formula(256, column, f"{cell(254, column)}*{cell(255, column)}")
    set_formula(256, "N", f"SUM({cell(256, 'C')}:{cell(256, 'M')})")
    set_formula(266, "C", safe_div(cell(248, "M"), cell(258, "C")))
    set_formula(267, "C", cell(255, "M"), "0.0000")
    set_formula(268, "C", f"{cell(266, 'C')}*{cell(267, 'C')}")
    set_formula(272, "C", f"{cell(256, 'N')}+{cell(268, 'C')}")
    set_formula(273, "C", cell(256, "N"))
    set_formula(274, "C", cell(268, "C"))
    set_formula(275, "C", f"{cell(142, 'C')}-{cell(137, 'C')}")
    set_formula(276, "C", f"{cell(272, 'C')}+{cell(275, 'C')}")

    cash_row = sources["cash_row"]
    set_formula(281, "D", source("assets", "C", cash_row))
    set_formula(282, "D", cell(254, "N"))
    set_formula(284, "D", f"({cell(142, 'C')}-{cell(137, 'C')})*0.95")
    set_formula(285, "D", f"-{cell(146, 'C')}")
    set_formula(283, "D", f"{cell(284, 'D')}+{cell(285, 'D')}")
    set_formula(286, "D", f"SUM({cell(281, 'D')}:{cell(283, 'D')})")

    going_debt_sources = {
        287: 75,
        288: 76,
        289: 77,
        290: 71,
        291: None,
        292: 72,
        293: 73,
        294: 74,
    }
    for row_number_value, source_row in going_debt_sources.items():
        set_formula(row_number_value, "C", cell(source_row, "C") if source_row else "0")
    set_formula(287, "D", f"MAX(MIN({cell(287, 'C')},{cell(286, 'D')}),0)")
    set_value(288, "D", 0)
    set_formula(289, "D", f"MAX(MIN({cell(289, 'C')},{cell(286, 'D')}-{cell(287, 'D')}),0)")
    set_formula(290, "D", f"MAX(MIN({cell(286, 'D')}-{cell(287, 'D')}-{cell(289, 'D')},MAX({cell(290, 'C')},{cell(192, 'G')})),0)")
    set_formula(291, "D", f"MIN({cell(286, 'D')}-{cell(287, 'D')}-{cell(289, 'D')}-{cell(290, 'D')},{cell(290, 'D')}*{cell(300, 'C')}/365*{cell(301, 'C')})")
    unsecured_denominator = f"{cell(292, 'C')}+{cell(293, 'C')}"
    unsecured_remainder = f"{cell(286, 'D')}-{cell(287, 'D')}-{cell(289, 'D')}-{cell(290, 'D')}-{cell(291, 'D')}"
    for row_number_value in (292, 293):
        set_formula(
            row_number_value,
            "D",
            f"MAX(MIN({cell(row_number_value, 'C')},({unsecured_remainder})*IFERROR({cell(row_number_value, 'C')}/({unsecured_denominator}),0)),0)",
        )
    set_value(294, "D", 0)
    for row_number_value in range(287, 295):
        set_formula(row_number_value, "E", safe_div(cell(row_number_value, "D"), cell(row_number_value, "C")), "0.00%")
    for column in ("C", "D"):
        set_formula(295, column, f"SUM({cell(287, column)}:{cell(294, column)})")
    set_formula(296, "D", f"{cell(286, 'D')}-{cell(295, 'D')}")
    set_formula(300, "C", f"{cell(164, 'C')}+{cell(165, 'C')}")
    set_formula(301, "C", f"{cell(163, 'C')}+0.005", "0.00%")

    set_formula(320, "C", cell(315, "E"))
    set_formula(320, "D", cell(276, "C"))
    set_formula(320, "E", f"{cell(320, 'D')}-{cell(320, 'C')}")

    comparison_sources = {
        326: (189, 287),
        327: (191, 289),
        328: (192, 290),
        329: (193, 292),
        330: (194, 293),
        331: (195, 294),
    }
    for row_number_value, (liquidation_row, going_row) in comparison_sources.items():
        set_formula(row_number_value, "C", cell(liquidation_row, "C"))
        set_formula(row_number_value, "D", cell(liquidation_row, "G"))
        set_formula(row_number_value, "E", cell(liquidation_row, "I"), "0.00%")
        set_formula(row_number_value, "F", cell(going_row, "D"))
        set_formula(row_number_value, "G", cell(going_row, "E"), "0.00%")
        set_formula(row_number_value, "H", f"IFERROR({cell(row_number_value, 'F')}-{cell(row_number_value, 'D')},0)")
        set_formula(row_number_value, "I", f"IFERROR({cell(row_number_value, 'G')}-{cell(row_number_value, 'E')},0)", "0.00%")
    for column in ("C", "D", "F", "H"):
        set_formula(332, column, f"SUM({cell(326, column)}:{cell(331, column)})")
    set_formula(337, "C", f'IF({cell(320, "D")}>{cell(320, "C")},"Positive","Negative")', None)
    set_formula(
        342,
        "C",
        f'IFERROR(IF(AND({cell(326, "I")}>=0,{cell(327, "I")}>=0,{cell(328, "I")}>=0,{cell(329, "I")}>=0,{cell(330, "I")}>=0),"Positive","Negative"),"Negative")',
        None,
    )
    set_formula(
        347,
        "C",
        f'IF(AND({cell(337, "C")}="Positive",{cell(342, "C")}="Positive",{cell(329, "G")}>=30%),"Positive","Negative")',
        None,
    )
    set_formula(352, "C", cell(347, "C"), None)

    return formulas, formats


def append_worksheet_review_sheet(
    workbook: Workbook,
    case: dict[str, Any],
    result: dict[str, Any],
    sheet_titles: dict[str, str],
    used_titles: set[str],
) -> None:
    review = result.get("worksheet_review", {})
    columns = ["행", "작업시트 항목"] + list(review.get("columns", []))
    review_rows = list(review.get("rows", []))
    row_index_by_number = {
        row.get("row_number"): index
        for index, row in enumerate(review_rows, start=2)
        if isinstance(row.get("row_number"), int)
    }
    formulas, formats = worksheet_review_formulas(case, result, sheet_titles, row_index_by_number)
    rows = []
    for row in review_rows:
        row_number_value = row.get("row_number", "")
        values = []
        for column, value in zip(REVIEW_COLUMNS, row.get("values", [])):
            values.append(formulas.get((row_number_value, column), value))
        rows.append([row_number_value, row.get("label", "")] + values)

    worksheet = append_table_sheet(workbook, "작업시트 확인", columns, rows, used_titles)
    for (row_number_value, column), number_format in formats.items():
        row_index = row_index_by_number.get(row_number_value)
        if row_index:
            worksheet[f"{column}{row_index}"].number_format = number_format


def build_extracted_workbook_download(case: dict[str, Any]) -> io.BytesIO:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    used_titles: set[str] = set()

    extracted_sheets = case.get("workbook", {}).get("sheets", [])
    if not extracted_sheets:
        worksheet = workbook.create_sheet("추출 데이터")
        worksheet["A1"] = "추출된 시트 데이터가 없습니다."
        style_extracted_worksheet(worksheet)
    else:
        for index, sheet in enumerate(extracted_sheets, start=1):
            title = safe_excel_sheet_title(sheet.get("document_label") or sheet.get("name") or f"Sheet{index}", used_titles)
            worksheet = workbook.create_sheet(title)
            pending_merges: list[tuple[int, int, int, int]] = []

            for row_number, row in enumerate(sheet.get("rows", []), start=1):
                for cell in row.get("cells", []):
                    source_col = int(cell.get("source_col", 0))
                    column_number = source_col + 1
                    worksheet.cell(row=row_number, column=column_number).value = extracted_excel_value(cell)

                    rowspan = max(1, int(cell.get("rowspan", 1) or 1))
                    colspan = max(1, int(cell.get("colspan", 1) or 1))
                    if rowspan > 1 or colspan > 1:
                        pending_merges.append((row_number, column_number, row_number + rowspan - 1, column_number + colspan - 1))

            for start_row, start_col, end_row, end_col in pending_merges:
                try:
                    worksheet.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
                except ValueError:
                    continue

            style_extracted_worksheet(worksheet)

    sheet_titles = append_case_input_sheets(workbook, case, used_titles)

    try:
        result = calculate_case_result(case)
    except Exception:
        result = None
    if result:
        append_result_sheets(workbook, result, used_titles)
        append_worksheet_review_sheet(workbook, case, result, sheet_titles, used_titles)

    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def case_progress(case: dict[str, Any]) -> dict[str, str]:
    if case.get("collateral_saved"):
        return {"label": "결과 확인", "endpoint": "result"}
    if case.get("income_saved"):
        return {"label": "담보 입력", "endpoint": "collateral"}
    if case.get("debt_saved"):
        return {"label": "손익추정", "endpoint": "income"}
    if case.get("financial_saved"):
        return {"label": "채무입력", "endpoint": "debt"}
    return {"label": "자산가치산정", "endpoint": "financial"}


def case_diagnosis_label(case: dict[str, Any]) -> str:
    if not case.get("collateral_saved"):
        return "입력 진행 중"
    try:
        diagnosis = calculate_case_result(case)["diagnosis"]
    except Exception:
        return "결과 확인 필요"
    return "회생 가능성 검토 대상" if diagnosis.get("overall_positive") else "추가 검토 필요"


def diagnosis_chip(label: str, positive: bool) -> dict[str, str]:
    return {
        "label": label,
        "status": "회생유리" if positive else "회생불리",
        "tone": "positive" if positive else "negative",
    }


def case_diagnosis_summary(case: dict[str, Any]) -> list[dict[str, str]]:
    if not case.get("collateral_saved"):
        return [
            {
                "label": "진행",
                "status": "입력중",
                "tone": "neutral",
            }
        ]

    try:
        result = calculate_case_result(case)
        diagnosis = result["diagnosis"]
    except Exception:
        return [
            {
                "label": "진단",
                "status": "확인필요",
                "tone": "warning",
            }
        ]

    overall_positive = bool(diagnosis.get("overall_positive"))
    return [
        diagnosis_chip("가치 비교", bool(diagnosis.get("value_positive"))),
        diagnosis_chip("변제율 비교", bool(diagnosis.get("repayment_positive"))),
        diagnosis_chip("채권자 동의 가능성", bool(diagnosis.get("consent_positive"))),
        {
            "label": "종합 결론",
            "status": "회생유리" if overall_positive else "추가검토",
            "tone": "positive" if overall_positive else "warning",
        },
    ]


def case_listing_for_user(user_id: str) -> list[dict[str, Any]]:
    listings = []
    for case in CASE_STORE.values():
        if case.get("user_id") != user_id:
            continue
        if case.get("deleted_at"):
            continue
        progress = case_progress(case)
        listings.append(
            {
                "case": case,
                "progress": progress,
                "diagnosis": case_diagnosis_label(case),
                "summary": case_diagnosis_summary(case),
            }
        )
    return sorted(
        listings,
        key=lambda item: item["case"].get("updated_at") or item["case"].get("created_at") or "",
        reverse=True,
    )


def admin_status_value(case: dict[str, Any]) -> str:
    status = str(case.get("admin_status") or "new")
    return status if status in ADMIN_STATUS_LABELS else "new"


def admin_status_label(case: dict[str, Any]) -> str:
    return ADMIN_STATUS_LABELS[admin_status_value(case)]


def admin_status_tone(case: dict[str, Any]) -> str:
    return ADMIN_STATUS_TONES.get(admin_status_value(case), "neutral")


def admin_case_contact(case: dict[str, Any], user: dict[str, Any] | None = None) -> dict[str, str]:
    consultation = case.get("consultation") if isinstance(case.get("consultation"), dict) else {}
    return {
        "company": (
            consultation.get("company")
            or case.get("company_name")
            or (user or {}).get("company")
            or "회사명 없음"
        ),
        "contact_name": consultation.get("contact_name") or (user or {}).get("contact_name", ""),
        "phone": consultation.get("phone") or (user or {}).get("phone", ""),
        "email": consultation.get("email") or (user or {}).get("email", ""),
        "message": consultation.get("message", ""),
    }


def admin_case_result_summary(case: dict[str, Any]) -> dict[str, Any] | None:
    if not case.get("collateral_saved"):
        return None
    try:
        result = calculate_case_result(case)
    except Exception:
        return None
    return {
        "summary": result.get("summary", {}),
        "diagnosis": result.get("diagnosis", {}),
        "comparison_rows": result.get("comparison_rows", []),
    }


def admin_case_listing() -> list[dict[str, Any]]:
    listings = []
    for case in CASE_STORE.values():
        if case.get("deleted_at"):
            continue
        user = USERS.get(str(case.get("user_id")))
        result = admin_case_result_summary(case)
        listings.append(
            {
                "case": case,
                "user": user,
                "contact": admin_case_contact(case, user),
                "progress": case_progress(case),
                "diagnosis": case_diagnosis_label(case),
                "summary": case_diagnosis_summary(case),
                "result": result,
                "admin_status": admin_status_value(case),
                "admin_status_label": admin_status_label(case),
                "admin_status_tone": admin_status_tone(case),
                "consultation_submitted": bool(case.get("consultation_submitted")),
            }
        )
    return sorted(
        listings,
        key=lambda item: item["case"].get("updated_at") or item["case"].get("created_at") or "",
        reverse=True,
    )


def admin_dashboard_stats(items: list[dict[str, Any]]) -> dict[str, int]:
    return {
        "case_count": len(items),
        "consultation_count": sum(1 for item in items if item["consultation_submitted"]),
        "new_count": sum(1 for item in items if item["admin_status"] == "new"),
        "completed_count": sum(1 for item in items if item["case"].get("collateral_saved")),
        "user_count": len(USERS),
    }


def admin_user_listing() -> list[dict[str, Any]]:
    user_case_counts: dict[str, int] = {}
    last_case_at: dict[str, str] = {}
    for case in CASE_STORE.values():
        user_id = str(case.get("user_id") or "")
        if not user_id or case.get("deleted_at"):
            continue
        user_case_counts[user_id] = user_case_counts.get(user_id, 0) + 1
        timestamp = str(case.get("updated_at") or case.get("created_at") or "")
        if timestamp > last_case_at.get(user_id, ""):
            last_case_at[user_id] = timestamp

    users = []
    for user in USERS.values():
        user_id = str(user.get("user_id"))
        users.append(
            {
                "user": user,
                "case_count": user_case_counts.get(user_id, 0),
                "last_case_at": last_case_at.get(user_id, ""),
            }
        )
    return sorted(users, key=lambda item: item["last_case_at"] or item["user"].get("created_at", ""), reverse=True)


def copy_case_for_user(source_case: dict[str, Any], user_id: str) -> dict[str, Any]:
    copied_case = deepcopy(source_case)
    original_id = source_case["case_id"]
    new_case_id = uuid.uuid4().hex[:12]
    created_at = now_iso()
    copied_case["case_id"] = new_case_id
    copied_case["source_case_id"] = original_id
    copied_case["user_id"] = user_id
    copied_case["created_at"] = created_at
    copied_case["updated_at"] = created_at
    copied_case["scenario_name"] = f"{source_case.get('scenario_name') or source_case.get('company_name', '분석')} 복사본"
    copied_case.pop("consultation", None)
    copied_case["consultation_submitted"] = False
    CASE_STORE[new_case_id] = copied_case
    remember_session_case(new_case_id)
    save_case(copied_case)
    return copied_case


def clean_url_values(values: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in values.items() if value not in (None, "")}


def mypage_message_from_request() -> str | None:
    messages = [
        ("joined", "회원가입이 완료되었습니다. 분석 이력을 이곳에서 확인할 수 있습니다."),
        ("deleted", "분석 이력을 삭제했습니다."),
        ("account_updated", "회원 정보를 수정했습니다."),
        ("email_verified", "이메일 인증이 완료되었습니다."),
        ("email_sent", "인증 메일을 발송했습니다."),
        ("email_not_sent", "인증 메일을 발송하지 못했습니다. AWS SES 설정을 확인해 주세요."),
        ("copied", "복사본을 생성했습니다."),
        ("renamed", "분석 이름을 저장했습니다."),
    ]
    for key, message in messages:
        if request.args.get(key) == "1":
            return message
    return None


@app.get("/")
def index():
    return render_template("landing.html", active_page="home")


@app.route("/signup", methods=["GET", "POST"])
def signup():
    if current_user():
        return redirect(url_for("mypage"))

    case_id = request.values.get("case_id", "")
    prefill_case = get_accessible_case(case_id) if case_id else None
    prefill_consultation = (prefill_case or {}).get("consultation", {})
    form = {
        "company": request.form.get("company", prefill_consultation.get("company", (prefill_case or {}).get("company_name", ""))).strip(),
        "contact_name": request.form.get("contact_name", prefill_consultation.get("contact_name", "")).strip(),
        "phone": request.form.get("phone", prefill_consultation.get("phone", "")).strip(),
        "email": normalize_email(request.form.get("email", prefill_consultation.get("email", ""))),
    }
    error = None

    if request.method == "POST":
        password = request.form.get("password", "")
        password_confirm = request.form.get("password_confirm", "")

        if not form["company"]:
            error = "회사명을 입력해 주세요."
        elif not form["contact_name"]:
            error = "담당자명을 입력해 주세요."
        elif not form["phone"]:
            error = "연락처를 입력해 주세요."
        elif not form["email"]:
            error = "이메일을 입력해 주세요."
        elif not EMAIL_RE.match(form["email"]):
            error = "이메일 형식을 확인해 주세요."
        elif find_user_by_email(form["email"]):
            error = "이미 가입된 이메일입니다."
        elif len(password) < 8:
            error = "비밀번호는 8자 이상 입력해 주세요."
        elif password != password_confirm:
            error = "비밀번호 확인이 일치하지 않습니다."
        elif request.form.get("privacy_consent") != "on":
            error = "개인정보 수집 및 이용에 동의해 주세요."
        else:
            user_id = uuid.uuid4().hex[:12]
            user = {
                "user_id": user_id,
                "company": form["company"],
                "contact_name": form["contact_name"],
                "phone": form["phone"],
                "email": form["email"],
                "password_hash": generate_password_hash(password),
                "email_verified": False,
                "created_at": now_iso(),
            }
            USERS[user_id] = user
            save_users()
            login_user(user)
            sent, _ = send_email_verification(user)
            save_users()
            return redirect(
                url_for(
                    "mypage",
                    **clean_url_values(
                        {
                            "joined": "1",
                            "email_sent": "1" if sent else "",
                            "email_not_sent": "" if sent else "1",
                        }
                    ),
                )
            )

    return render_template(
        "signup.html",
        active_page="signup",
        form=form,
        case_id=case_id,
        error=error,
    )


@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user():
        return redirect(url_for("mypage"))

    email = normalize_email(request.form.get("email"))
    error = None
    message = "회원탈퇴가 완료되었습니다." if request.args.get("account_deleted") == "1" else None
    if request.method == "POST":
        user = find_user_by_email(email)
        password = request.form.get("password", "")
        if not user or not check_password_hash(user.get("password_hash", ""), password):
            error = "이메일 또는 비밀번호를 확인해 주세요."
        else:
            login_user(user)
            next_path = request.args.get("next") or request.form.get("next") or url_for("mypage")
            if not next_path.startswith("/"):
                next_path = url_for("mypage")
            return redirect(next_path)

    return render_template(
        "login.html",
        active_page="login",
        email=email,
        next_path=request.args.get("next", ""),
        error=error,
        message=message,
    )


@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    user = current_user()
    assert user is not None
    form = {
        "company": request.form.get("company", user.get("company", "")).strip(),
        "contact_name": request.form.get("contact_name", user.get("contact_name", "")).strip(),
        "phone": request.form.get("phone", user.get("phone", "")).strip(),
        "email": normalize_email(request.form.get("email", user.get("email", ""))),
    }
    error = None

    if request.method == "POST":
        email_owner = find_user_by_email(form["email"]) if form["email"] else None
        email_changed = form["email"] != normalize_email(user.get("email"))
        if not form["company"]:
            error = "회사명을 입력해 주세요."
        elif not form["contact_name"]:
            error = "담당자명을 입력해 주세요."
        elif not form["phone"]:
            error = "연락처를 입력해 주세요."
        elif not form["email"]:
            error = "이메일을 입력해 주세요."
        elif not EMAIL_RE.match(form["email"]):
            error = "이메일 형식을 확인해 주세요."
        elif email_owner and email_owner.get("user_id") != user.get("user_id"):
            error = "이미 가입된 이메일입니다."
        else:
            user["company"] = form["company"]
            user["contact_name"] = form["contact_name"]
            user["phone"] = form["phone"]
            user["email"] = form["email"]
            if email_changed:
                user["email_verified"] = False
                user.pop("email_verified_at", None)
                sent, _ = send_email_verification(user)
            else:
                sent = False
            user["updated_at"] = now_iso()
            save_users()
            return redirect(
                url_for(
                    "mypage",
                    **clean_url_values(
                        {
                            "account_updated": "1",
                            "email_sent": "1" if email_changed and sent else "",
                            "email_not_sent": "1" if email_changed and not sent else "",
                        }
                    ),
                )
            )

    return render_template(
        "profile.html",
        active_page="mypage",
        form=form,
        error=error,
        message=(
            "인증 메일을 발송했습니다."
            if request.args.get("email_sent") == "1"
            else (
                "이메일 인증이 이미 완료되어 있습니다."
                if request.args.get("email_already_verified") == "1"
                else None
            )
        ),
        warning=(
            "인증 메일을 발송하지 못했습니다. AWS SES 설정을 확인해 주세요."
            if request.args.get("email_not_sent") == "1"
            else None
        ),
    )


@app.post("/email-verification/send")
@login_required
def send_email_verification_route():
    user = current_user()
    assert user is not None
    if user.get("email_verified"):
        return redirect(url_for("profile", email_already_verified="1"))
    sent, _ = send_email_verification(user)
    save_users()
    return redirect(
        url_for(
            "profile",
            **clean_url_values(
                {
                    "email_sent": "1" if sent else "",
                    "email_not_sent": "" if sent else "1",
                }
            ),
        )
    )


@app.get("/verify-email")
def verify_email():
    token = request.args.get("token", "")
    status = "invalid"
    title = "이메일 인증 실패"
    message = "인증 링크가 올바르지 않습니다. 다시 인증 메일을 요청해 주세요."
    try:
        payload = load_email_verification_token(token)
    except SignatureExpired:
        status = "expired"
        message = "인증 링크 유효기간이 만료되었습니다. 다시 인증 메일을 요청해 주세요."
    except BadSignature:
        message = "인증 링크를 확인할 수 없습니다. 다시 인증 메일을 요청해 주세요."
    else:
        user = USERS.get(str(payload.get("user_id")))
        email = normalize_email(payload.get("email"))
        if user and normalize_email(user.get("email")) == email:
            user["email_verified"] = True
            user["email_verified_at"] = now_iso()
            save_users()
            if current_user_id() == user.get("user_id"):
                return redirect(url_for("mypage", email_verified="1"))
            status = "success"
            title = "이메일 인증 완료"
            message = "이메일 인증이 완료되었습니다. 이제 로그인하여 분석 이력을 확인할 수 있습니다."

    return render_template(
        "email_verification_result.html",
        active_page="login",
        status=status,
        title=title,
        message=message,
    )


@app.route("/account/delete", methods=["GET", "POST"])
@login_required
def delete_account():
    user = current_user()
    assert user is not None
    error = None

    if request.method == "POST":
        email = normalize_email(request.form.get("email"))
        if email != normalize_email(user.get("email")):
            error = "현재 이메일을 정확히 입력해 주세요."
        elif request.form.get("confirm_delete") != "on":
            error = "회원탈퇴 확인에 체크해 주세요."
        else:
            user_id = user["user_id"]
            for case in CASE_STORE.values():
                if case.get("user_id") == user_id and not case.get("deleted_at"):
                    case["deleted_at"] = now_iso()
                    save_case(case)
            USERS.pop(user_id, None)
            save_users()
            session.clear()
            return redirect(url_for("login", account_deleted="1"))

    return render_template(
        "delete_account.html",
        active_page="mypage",
        error=error,
    )


@app.post("/logout")
def logout():
    session.pop("user_id", None)
    return redirect(url_for("index"))


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if current_admin():
        return redirect(url_for("admin_dashboard"))

    next_path = request.values.get("next", "")
    if not str(next_path).startswith("/admin"):
        next_path = url_for("admin_dashboard")

    error = None
    username = request.form.get("username", admin_username()).strip()

    if request.method == "POST":
        if not admin_password_is_configured():
            error = "관리자 비밀번호 환경변수가 설정되어 있지 않습니다."
        elif username != admin_username() or not admin_password_matches(request.form.get("password", "")):
            error = "관리자 아이디 또는 비밀번호를 확인해 주세요."
        else:
            session["admin_user"] = username
            return redirect(next_path)

    return render_template(
        "admin_login.html",
        active_page="admin",
        username=username,
        next_path=next_path,
        error=error,
        is_configured=admin_password_is_configured(),
    )


@app.post("/admin/logout")
def admin_logout():
    session.pop("admin_user", None)
    return redirect(url_for("admin_login"))


@app.get("/admin")
@admin_login_required
def admin_dashboard():
    items = admin_case_listing()
    return render_template(
        "admin_dashboard.html",
        active_page="admin",
        items=items,
        stats=admin_dashboard_stats(items),
        status_options=ADMIN_STATUS_OPTIONS,
    )


@app.get("/admin/users")
@admin_login_required
def admin_users():
    return render_template(
        "admin_users.html",
        active_page="admin",
        users=admin_user_listing(),
    )


@app.get("/admin/cases/<case_id>/extracted.xlsx")
@admin_login_required
def admin_case_extracted_workbook(case_id: str):
    case = CASE_STORE.get(case_id)
    if not case or case.get("deleted_at"):
        return redirect(url_for("admin_dashboard"))

    output = build_extracted_workbook_download(case)
    company_name = safe_download_basename(case.get("scenario_name") or case.get("company_name") or case_id)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"{company_name}_검증데이터.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/cases/<case_id>", methods=["GET", "POST"])
@admin_login_required
def admin_case_detail(case_id: str):
    case = CASE_STORE.get(case_id)
    if not case or case.get("deleted_at"):
        return redirect(url_for("admin_dashboard"))

    if request.method == "POST":
        status = request.form.get("admin_status", "new")
        if status not in ADMIN_STATUS_LABELS:
            status = "new"
        case["admin_status"] = status
        case["admin_assignee"] = clean_contact_text(request.form.get("admin_assignee"), 80)
        case["admin_memo"] = str(request.form.get("admin_memo") or "").strip()[:3000]
        case["admin_updated_at"] = now_iso()
        write_case(case)
        return redirect(url_for("admin_case_detail", case_id=case_id, saved="1"))

    user = USERS.get(str(case.get("user_id")))
    result = admin_case_result_summary(case)
    return render_template(
        "admin_case_detail.html",
        active_page="admin",
        case=case,
        user=user,
        contact=admin_case_contact(case, user),
        progress=case_progress(case),
        diagnosis=case_diagnosis_label(case),
        summary=case_diagnosis_summary(case),
        result=result,
        status_options=ADMIN_STATUS_OPTIONS,
        admin_status=admin_status_value(case),
        admin_status_label=admin_status_label(case),
        admin_status_tone=admin_status_tone(case),
        message="관리자 메모가 저장되었습니다." if request.args.get("saved") == "1" else None,
    )


@app.get("/mypage")
@login_required
def mypage():
    user = current_user()
    assert user is not None
    return render_template(
        "mypage.html",
        active_page="mypage",
        cases=case_listing_for_user(user["user_id"]),
        message=mypage_message_from_request(),
    )


@app.post("/cases/<case_id>/copy")
@login_required
def copy_case(case_id: str):
    user = current_user()
    source_case = get_accessible_case(case_id)
    if user is None or source_case is None:
        return redirect(url_for("mypage"))
    copied_case = copy_case_for_user(source_case, user["user_id"])
    return redirect(url_for("mypage", copied="1", copied_id=copied_case["case_id"]))


@app.post("/cases/<case_id>/rename")
@login_required
def rename_case(case_id: str):
    case = get_accessible_case(case_id)
    if case is None:
        return redirect(url_for("mypage"))
    scenario_name = request.form.get("scenario_name", "").strip()
    if scenario_name == str(case.get("company_name", "")).strip():
        scenario_name = ""
    case["scenario_name"] = scenario_name
    save_case(case)
    if request.headers.get("X-Requested-With") == "fetch":
        return ("", 204)
    return redirect(url_for("mypage", renamed="1"))


@app.post("/cases/<case_id>/delete")
@login_required
def delete_case(case_id: str):
    case = get_accessible_case(case_id)
    if case is None:
        return redirect(url_for("mypage"))
    case["deleted_at"] = now_iso()
    save_case(case)
    return redirect(url_for("mypage", deleted="1"))


@app.get("/analysis")
def analysis_index():
    return render_template("index.html", workbook=None, case=None, error=None)


@app.get("/upload")
def upload_form():
    return redirect(url_for("analysis_index"))


@app.get("/upload/<case_id>")
def upload_preview(case_id: str):
    case = get_accessible_case(case_id)
    if case is None:
        return render_template(
            "index.html",
            workbook=None,
            case=None,
            error="작업 정보를 찾을 수 없습니다. 파일을 다시 업로드해 주세요.",
        ), 404

    return render_template(
        "index.html",
        workbook=case["workbook"],
        case=case,
        error=None,
    )


@app.post("/upload")
def upload():
    try:
        workbook = build_document_workbook(request.files)
        case = create_case(workbook)
    except Exception as exc:
        return render_template(
            "index.html",
            workbook=None,
            case=None,
            error=f"엑셀을 읽는 중 오류가 발생했습니다: {exc}",
        ), 400

    return redirect(url_for("financial", case_id=case["case_id"], uploaded="1"))


@app.get("/financial/<case_id>")
def financial(case_id: str):
    case = get_accessible_case(case_id)
    if case is None:
        return render_template(
            "financial.html",
            case=None,
            rows=[],
            error="작업 정보를 찾을 수 없습니다. 파일을 다시 업로드해 주세요.",
            message=None,
        ), 404

    return render_template(
        "financial.html",
        case=case,
        rows=case["financial_rows"],
        error=None,
        message=(
            "파일이 업로드되었습니다. 자산가치산정을 확인해 주세요."
            if request.args.get("uploaded") == "1"
            else (
                "기존 분석을 복사했습니다. 필요한 숫자를 수정해 다시 분석할 수 있습니다."
                if request.args.get("copied") == "1"
                else None
            )
        ),
    )


@app.post("/financial/<case_id>/save")
def save_financial(case_id: str):
    case = get_accessible_case(case_id)
    if case is None:
        return redirect(url_for("analysis_index"))

    rows = case["financial_rows"]
    errors: list[str] = []

    for index, row in enumerate(rows):
        if not row["is_editable"]:
            continue

        audit_key = f"audit_value_{index}"
        liquidation_key = f"liquidation_value_{index}"
        audit_raw = request.form.get(audit_key, row["audit_value"])
        liquidation_raw = request.form.get(liquidation_key, row["liquidation_value"])

        audit_display, audit_number = clean_submitted_number(audit_raw)
        liquidation_display, liquidation_number = clean_submitted_number(liquidation_raw)

        if audit_number is not None and liquidation_number is not None and liquidation_number > audit_number:
            errors.append(f"{row['account']}: 청산가치는 실사가치보다 클 수 없습니다.")

        row["audit_value"] = audit_display
        row["liquidation_value"] = liquidation_display

    if errors:
        return render_template(
            "financial.html",
            case=case,
            rows=rows,
            error=" ".join(errors),
            message=None,
        ), 400

    case["financial_saved"] = True
    save_case(case)
    if request.form.get("next") == "debt":
        return redirect(url_for("debt", case_id=case_id))

    return render_template(
        "financial.html",
        case=case,
        rows=rows,
        error=None,
        message="재무 데이터가 저장되었습니다.",
    )


@app.get("/debt/<case_id>")
def debt(case_id: str):
    case = get_accessible_case(case_id)
    if case is None:
        return render_template(
            "debt.html",
            case=None,
            rows=[],
            error="작업 정보를 찾을 수 없습니다. 파일을 다시 업로드해 주세요.",
            message=None,
        ), 404

    return render_template(
        "debt.html",
        case=case,
        rows=case["debt_rows"],
        error=None,
        message=None,
    )


@app.post("/debt/<case_id>/save")
def save_debt(case_id: str):
    case = get_accessible_case(case_id)
    if case is None:
        return redirect(url_for("analysis_index"))

    rows = case["debt_rows"]
    for index, row in enumerate(rows):
        raw_amount = request.form.get(f"debt_amount_{index}", row["debt_amount"])
        amount_display, _ = clean_submitted_number(raw_amount)
        row["debt_amount"] = amount_display
        row["collateral_type"] = request.form.get(f"collateral_type_{index}", row.get("collateral_type", ""))
        row["audit_value"] = request.form.get(f"audit_value_{index}", row.get("audit_value", ""))

    case["debt_saved"] = True
    save_case(case)
    if request.form.get("next") == "income":
        return redirect(url_for("income", case_id=case_id))

    return render_template(
        "debt.html",
        case=case,
        rows=rows,
        error=None,
        message="부채 데이터가 저장되었습니다. 다음 단계는 손익추정입니다.",
    )


@app.get("/income/<case_id>")
def income(case_id: str):
    case = get_accessible_case(case_id)
    if case is None:
        return render_template(
            "income.html",
            case=None,
            sales_rows=[],
            expense_rows=[],
            error="작업 정보를 찾을 수 없습니다. 파일을 다시 업로드해 주세요.",
            message=None,
        ), 404

    rows = case["income_rows"]
    return render_template(
        "income.html",
        case=case,
        sales_rows=[row for row in rows if row["section"] == "sales"],
        expense_rows=[row for row in rows if row["section"] == "expense"],
        error=None,
        message=None,
    )


@app.post("/income/<case_id>/save")
def save_income(case_id: str):
    case = get_accessible_case(case_id)
    if case is None:
        return redirect(url_for("analysis_index"))

    rows = case["income_rows"]
    for index, row in enumerate(rows):
        if not row["is_editable"]:
            continue

        row["cost_type"] = request.form.get(f"cost_type_{index}", row.get("cost_type", ""))
        row["monthly_average_display"] = request.form.get(
            f"monthly_average_display_{index}",
            row.get("monthly_average_display", ""),
        )

        value_kind = request.form.get(f"value_kind_{index}", row.get("value_kind", "amount"))
        final_raw = request.form.get(f"final_value_{index}", row["final_value"])
        final_display, _ = clean_submitted_income_value(final_raw, value_kind)

        row["value_kind"] = value_kind
        row["final_value"] = final_display

    case["income_saved"] = True
    save_case(case)
    if request.form.get("next") == "collateral":
        return redirect(url_for("collateral", case_id=case_id))

    return render_template(
        "income.html",
        case=case,
        sales_rows=[row for row in rows if row["section"] == "sales"],
        expense_rows=[row for row in rows if row["section"] == "expense"],
        error=None,
        message="손익 데이터가 저장되었습니다. 다음 단계는 담보등자산입니다.",
    )


@app.get("/collateral/<case_id>")
def collateral(case_id: str):
    case = get_accessible_case(case_id)
    if case is None:
        return render_template(
            "collateral.html",
            case=None,
            collateral_rows=[],
            rent_rows=[],
            error="작업 정보를 찾을 수 없습니다. 파일을 다시 업로드해 주세요.",
            message=None,
        ), 404

    return render_template(
        "collateral.html",
        case=case,
        collateral_rows=case["collateral_rows"],
        rent_rows=case["rent_rows"],
        error=None,
        message=None,
    )


@app.post("/collateral/<case_id>/save")
def save_collateral(case_id: str):
    case = get_accessible_case(case_id)
    if case is None:
        return redirect(url_for("analysis_index"))

    collateral_rows = case["collateral_rows"]
    rent_rows = case["rent_rows"]
    errors: list[str] = []

    for index, row in enumerate(collateral_rows):
        audit_raw = request.form.get(f"audit_value_{index}", row["audit_value"])
        liquidation_raw = request.form.get(
            f"liquidation_value_{index}",
            row["liquidation_value"],
        )

        audit_display, audit_number = clean_submitted_number(audit_raw)
        liquidation_display, liquidation_number = clean_submitted_number(liquidation_raw)

        if audit_number is not None and liquidation_number is not None and liquidation_number > audit_number:
            errors.append(f"{row['category']}: 청산가치는 실사가치보다 클 수 없습니다.")

        row["audit_value"] = audit_display
        row["liquidation_value"] = liquidation_display

    for index, row in enumerate(rent_rows):
        amount_raw = request.form.get(f"rent_amount_{index}", row["amount"])
        amount_display, _ = clean_submitted_number(amount_raw)
        row["amount"] = amount_display

    if errors:
        return render_template(
            "collateral.html",
            case=case,
            collateral_rows=collateral_rows,
            rent_rows=rent_rows,
            error=" ".join(errors),
            message=None,
        ), 400

    case["collateral_saved"] = True
    save_case(case)
    if request.form.get("next") == "result":
        return redirect(url_for("result", case_id=case_id))

    return render_template(
        "collateral.html",
        case=case,
        collateral_rows=collateral_rows,
        rent_rows=rent_rows,
        error=None,
        message="담보등자산 데이터가 저장되었습니다. 다음 단계는 결과확인입니다.",
    )


@app.get("/result/<case_id>")
def result(case_id: str):
    case = get_accessible_case(case_id)
    if case is None:
        return render_template(
            "result.html",
            case=None,
            error="작업 정보를 찾을 수 없습니다. 파일을 다시 업로드해 주세요.",
        ), 404

    return render_template(
        "result.html",
        case=case,
        calculation_result=calculate_case_result(case),
        error=None,
        message="상담 신청이 접수되었습니다. 상세 분석 결과를 확인할 수 있습니다."
        if request.args.get("consultation") == "submitted"
        else None,
        consultation_form=case.get("consultation", {}),
    )


@app.post("/consultation/<case_id>")
def submit_consultation(case_id: str):
    case = get_accessible_case(case_id)
    if case is None:
        return redirect(url_for("analysis_index"))

    form = consultation_form_from_request()
    errors = validate_consultation_form(form)
    result_data = calculate_case_result(case)

    if errors:
        return render_template(
            "result.html",
            case=case,
            calculation_result=result_data,
            error=" ".join(errors),
            message=None,
            consultation_form=form,
        ), 400

    append_consultation_log(case, form, result_data)
    case["consultation"] = form
    case["consultation_submitted"] = True
    save_case(case)
    return redirect(url_for("result", case_id=case_id, consultation="submitted"))


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=True)
